"""Lette tester for Excel- og PDF-eksport: gyldige filer med forventede ark."""

import io

import openpyxl
import pandas as pd
import pytest

from ryfast_app.config import COMPARE_YEARS
from ryfast_app.exports import excel as excel_mod
from ryfast_app.exports import pdf as pdf_mod
from tests.fixtures import two_point_year


@pytest.fixture
def comparison_df():
    return pd.DataFrame(
        {
            "Month": list(range(1, 13)),
            "Month Name": [
                "Januar", "Februar", "Mars", "April", "Mai", "Juni",
                "Juli", "August", "September", "Oktober", "November", "Desember",
            ],
            "2024": [1500.0] * 12,
            "2025": [1600.0] * 12,
        }
    )


@pytest.fixture
def _uten_api(monkeypatch):
    """Rapportbyggerne re-henter data; testene skal ikke treffe nettet."""
    monkeypatch.setattr(excel_mod, "fetch_batch_traffic_data", lambda *a, **k: two_point_year())
    monkeypatch.setattr(pdf_mod, "fetch_batch_traffic_data", lambda *a, **k: two_point_year())
    monkeypatch.setattr(pdf_mod, "_pdf_embed_figure", lambda pdf, fig: None)


class TestExportToExcel:
    def test_gyldig_arbeidsbok_med_ark(self, comparison_df):
        data = excel_mod.export_to_excel(comparison_df)
        wb = openpyxl.load_workbook(io.BytesIO(data))
        assert "Trafikkdata" in wb.sheetnames
        assert "Vekstrater" in wb.sheetnames  # to årskolonner -> vekstark
        assert "Sesongmønstre" in wb.sheetnames


class TestBuildExcelReport:
    def test_full_rapport_uten_nett(self, comparison_df, _uten_api):
        data = excel_mod.build_excel_report(
            comparison_df,
            point="Ryfast (sum tunneler)",
            comparison_mode=COMPARE_YEARS,
            year_list=[2024, 2025],
            year=2025,
            point_ids=["punkt_a", "punkt_b"],
            timeout_s=5,
            use_cache=False,
            coverage_threshold=50.0,
            ryfast_include_ramp=True,
        )
        wb = openpyxl.load_workbook(io.BytesIO(data))
        for sheet in ("Summary", "Metadata", "Data"):
            assert sheet in wb.sheetnames
        ws = wb["Data"]
        assert ws.cell(row=1, column=1).value == "Month"

    def test_delvis_aar_med_na_verdier(self, comparison_df, _uten_api):
        # Inneværende år har bare data t.o.m. juli; resten er NA (pd.NA via Int64)
        comparison_df["2025"] = comparison_df["2025"].where(comparison_df["Month"] <= 7).astype("Int64")
        data = excel_mod.build_excel_report(
            comparison_df,
            point="Ryfast (sum tunneler)",
            comparison_mode=COMPARE_YEARS,
            year_list=[2024, 2025],
            year=2025,
            point_ids=["punkt_a", "punkt_b"],
            timeout_s=5,
            use_cache=False,
            coverage_threshold=50.0,
            ryfast_include_ramp=True,
        )
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb["Data"]
        assert ws.cell(row=9, column=4).value is None  # august 2025 -> tom celle


class TestBuildPdfReport:
    def test_rapport_uten_figurer(self, comparison_df, _uten_api):
        data = pdf_mod.build_pdf_report(
            comparison_df,
            point="Ryfast (sum tunneler)",
            comparison_mode=COMPARE_YEARS,
            year_list=[2024, 2025],
            year=2025,
            point_ids=["punkt_a", "punkt_b"],
            timeout_s=5,
            use_cache=False,
            coverage_threshold=50.0,
            ryfast_include_ramp=True,
        )
        assert isinstance(data, (bytes, bytearray))
        assert bytes(data).startswith(b"%PDF")
