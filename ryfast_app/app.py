import calendar
import io
import logging
import os
import tempfile
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
import altair as alt
import plotly.express as px
import plotly.graph_objects as go
import requests
import streamlit as st
from fpdf import FPDF

try:
    import openpyxl  # noqa: F401  # pyright: ignore[reportUnusedImport]  # pylint: disable=unused-import
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Alignment, Font, PatternFill

    OPENPYXL_AVAILABLE = True
except Exception:
    openpyxl = None  # type: ignore[assignment]
    Workbook = None  # type: ignore[assignment]
    BarChart = None  # type: ignore[assignment]
    LineChart = None  # type: ignore[assignment]
    Reference = None  # type: ignore[assignment]
    CellIsRule = None  # type: ignore[assignment]
    Alignment = None  # type: ignore[assignment]
    Font = None  # type: ignore[assignment]
    PatternFill = None  # type: ignore[assignment]
    OPENPYXL_AVAILABLE = False


logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

URL = "https://trafikkdata-api.atlas.vegvesen.no"

TRAFFIC_POINTS = {
    "Ryfast (sum tunneler)": {
        "ids": [
            "99040V2725982",
            "00911V2725983",
            "10239V2725979",
            "62464V2725991",
            "92743V2726085",
            "25926V2725990",
        ],
        "description": "Sum av Ryfylketunnelen + Hundvågtunnelen (Ryfast totalt)",
        "opened": "2019-12-30 / 2020-04-22",
    },
    "Ryfylketunnelen": {
        "ids": ["99040V2725982", "00911V2725983"],
        "description": "Ryfylketunnelen - hovedforbindelse til Ryfylke",
        "opened": "2019-12-30",
    },
    "Hundvågtunnelen": {
        "ids": ["10239V2725979", "62464V2725991", "92743V2726085", "25926V2725990"],
        "description": "Hundvågtunnelen - forbindelse til Hundvåg og Eiganes",
        "opened": "2020-04-22",
    },
    "Bybrua": {
        "ids": {"Mot nord": ["17949V320695"], "Mot sør": ["54184V320694"]},
        "description": "Bybrua - historisk broforbindelse over Strømsteinsundet",
        "opened": "Historisk",
    },
}

HUNDVAG_TUNNEL_IDS_UTEN_PÅRAMPE = ["10239V2725979", "92743V2726085"]
HUNDVAG_TUNNEL_RAMP_IDS = ["62464V2725991", "25926V2725990"]

MONTH_NAMES = [
    "Januar",
    "Februar",
    "Mars",
    "April",
    "Mai",
    "Juni",
    "Juli",
    "August",
    "September",
    "Oktober",
    "November",
    "Desember",
]

DEFAULT_YEARS = "2024,2025"
YEAR_RANGE = range(2019, 2027)
API_MAX_RETRIES = 3
API_RETRY_DELAY = 1
API_CACHE_TTL = 24 * 3600
FULL_COVERAGE_TOL_PCT = 0.05  # tolerance to avoid float noise around 100%
ANOMALY_THRESHOLD_PCT = 20.0

COMPARE_YEARS = "Sammenlign år"
COMPARE_MONTHS = "Sammenlign måneder"
COMPARE_WEEKS = "Sammenlign uker"

QUERY_TEMPLATE = """
query {{
  trafficData(trafficRegistrationPointId: "{point_id}") {{
    volume {{
      average {{
        daily {{
          byMonth(year: {year}) {{
            month
            total {{
              volume {{
                average
                confidenceInterval {{
                  lowerBound
                  upperBound
                }}
              }}
              coverage {{
                percentage
              }}
            }}
          }}
        }}
      }}
    }}
  }}
}}
"""

WEEKLY_QUERY_TEMPLATE = """
query {{
  trafficData(trafficRegistrationPointId: "{point_id}") {{
    volume {{
      byDay(from: "{from_date}", to: "{to_date}") {{
        edges {{
          node {{
            from
            to
            total {{
              volumeNumbers {{
                volume
              }}
              coverage {{
                percentage
              }}
            }}
          }}
        }}
      }}
    }}
  }}
}}
"""

POINT_ID_LABELS: Dict[str, str] = {
    "99040V2725982": "Ryfylketunnelen (A)",
    "00911V2725983": "Ryfylketunnelen (B)",
    "10239V2725979": "Hundvågtunnelen (A)",
    "62464V2725991": "Hundvågtunnelen (pårampe?)",
    "92743V2726085": "Hundvågtunnelen (B)",
    "25926V2725990": "Hundvågtunnelen (pårampe?)",
    "17949V320695": "Bybrua (Mot nord)",
    "54184V320694": "Bybrua (Mot sør)",
}


def format_number(x):
    if pd.isna(x):
        return "N/A"
    if isinstance(x, (int, float, np.integer, np.floating)):
        if float(x).is_integer():
            return f"{int(x):,}".replace(",", " ")
        return f"{x:,.1f}".replace(",", " ")
    if isinstance(x, str):
        try:
            return format_number(float(x))
        except ValueError:
            return x
    return str(x)


def days_in_year(year: int) -> int:
    return 366 if calendar.isleap(year) else 365


def init_session_state():
    if "comparison_history" not in st.session_state:
        st.session_state.comparison_history = []
    if "last_result" not in st.session_state:
        st.session_state.last_result = None
    if "api_errors" not in st.session_state:
        st.session_state.api_errors = []


def record_api_error(message: str, query: Optional[str] = None) -> None:
    try:
        errors = list(st.session_state.get("api_errors", []))
        errors.append(
            {
                "ts": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "message": str(message),
                "query": (query[:600] + "…") if query and len(query) > 600 else query,
            }
        )
        st.session_state.api_errors = errors[-50:]
    except Exception as exc:
        logger.debug("record_api_error: klarte ikke oppdatere session_state: %s", exc)


def render_api_status_sidebar() -> None:
    errors = st.session_state.get("api_errors", []) or []
    with st.sidebar.expander("🧾 API-feil / status", expanded=False):
        if not errors:
            st.caption("Ingen registrerte API-feil i denne sesjonen.")
            return
        st.warning(f"{len(errors)} API-feil registrert i denne sesjonen.")
        if st.button("Tøm API-feil", type="secondary"):
            st.session_state.api_errors = []
            st.rerun()
        st.dataframe(pd.DataFrame(errors).iloc[::-1], use_container_width=True, hide_index=True)


def _fetch_data_uncached(query: str, timeout_s: int) -> Optional[Dict]:
    for attempt in range(API_MAX_RETRIES):
        try:
            response = requests.post(URL, json={"query": query}, timeout=timeout_s)
            response.raise_for_status()
            data = response.json()
            if "errors" in data:
                logger.error("GraphQL errors: %s", data["errors"])
                record_api_error(f"GraphQL error: {data['errors']}", query=query)
                return None
            return data
        except requests.Timeout:
            logger.warning("Timeout on attempt %s/%s", attempt + 1, API_MAX_RETRIES)
            record_api_error(f"Timeout (forsøk {attempt + 1}/{API_MAX_RETRIES})", query=query)
            if attempt == API_MAX_RETRIES - 1:
                return None
            time.sleep(API_RETRY_DELAY * (attempt + 1))
        except requests.RequestException as e:
            logger.warning("Request failed on attempt %s/%s: %s", attempt + 1, API_MAX_RETRIES, str(e))
            record_api_error(f"RequestException (forsøk {attempt + 1}/{API_MAX_RETRIES}): {e}", query=query)
            if attempt == API_MAX_RETRIES - 1:
                return None
            time.sleep(API_RETRY_DELAY * (attempt + 1))
    return None


@st.cache_data(ttl=API_CACHE_TTL, show_spinner=False)
def _fetch_data_cached(query: str, timeout_s: int) -> Optional[Dict]:
    return _fetch_data_uncached(query, timeout_s)


def fetch_data(query: str, timeout_s: int, use_cache: bool) -> Optional[Dict]:
    return _fetch_data_cached(query, timeout_s) if use_cache else _fetch_data_uncached(query, timeout_s)


def fetch_batch_traffic_data(point_ids: List[str], year: int, timeout_s: int, use_cache: bool) -> Dict[str, List[Dict]]:
    if year < 2019:
        return {}

    fetch_fn = _fetch_data_cached if use_cache else _fetch_data_uncached
    result: Dict[str, List[Dict]] = {}
    with ThreadPoolExecutor(max_workers=min(len(point_ids), 6)) as executor:
        future_to_point = {
            executor.submit(fetch_fn, QUERY_TEMPLATE.format(point_id=pid, year=year), timeout_s): pid
            for pid in point_ids
        }
        for future, point_id in future_to_point.items():
            try:
                data = future.result()
                if data and data.get("data", {}).get("trafficData"):
                    monthly = data["data"]["trafficData"]["volume"]["average"]["daily"]["byMonth"]
                    if monthly:
                        result[point_id] = monthly
            except Exception as e:
                logger.error("Feil ved henting av data for punkt %s, år %s: %s", point_id, year, str(e))

    return result


def fetch_weekly_traffic_data(
    point_ids: List[str], year: int, week_numbers: List[int], timeout_s: int, use_cache: bool
) -> Tuple[Dict[str, Dict[str, float]], Dict[str, Dict[str, float]]]:
    if year < 2019:
        return {}, {}

    # Pre-compute valid week date ranges (ISO 8601: Jan 4 is always in week 1)
    jan_4 = datetime(year, 1, 4)
    week_1_monday = jan_4 - timedelta(days=jan_4.isocalendar()[2] - 1)
    valid_weeks: List[Tuple[int, str, str]] = []
    for week_num in week_numbers:
        week_start = week_1_monday + timedelta(weeks=week_num - 1)
        week_end = week_start + timedelta(days=6)
        if week_start.isocalendar()[0] != year or week_end.isocalendar()[0] != year:
            continue
        valid_weeks.append((
            week_num,
            week_start.strftime("%Y-%m-%dT00:00:00+01:00"),
            week_end.strftime("%Y-%m-%dT23:59:59+01:00"),
        ))

    if not valid_weeks or not point_ids:
        return {}, {}

    result: Dict[str, Dict[str, float]] = {}
    cov_result: Dict[str, Dict[str, float]] = {}
    fetch_fn = _fetch_data_cached if use_cache else _fetch_data_uncached

    def _fetch_one(week_num: int, point_id: str, from_date: str, to_date: str):
        query = WEEKLY_QUERY_TEMPLATE.format(point_id=point_id, from_date=from_date, to_date=to_date)
        return week_num, point_id, fetch_fn(query, timeout_s)

    max_workers = min(len(valid_weeks) * len(point_ids), 12)
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {
            executor.submit(_fetch_one, week_num, point_id, from_date, to_date): (week_num, point_id)
            for week_num, from_date, to_date in valid_weeks
            for point_id in point_ids
        }
        for future in as_completed(futures):
            try:
                week_num, point_id, data = future.result()
                if data and data.get("data", {}).get("trafficData"):
                    edges = data["data"]["trafficData"]["volume"]["byDay"]["edges"] or []
                    total_volume = 0.0
                    valid_days = 0
                    cov_sum = 0.0
                    cov_days = 0
                    for edge in edges:
                        volume = edge["node"]["total"]["volumeNumbers"]["volume"]
                        if volume is not None:
                            total_volume += float(volume)
                            valid_days += 1
                        cov = edge["node"]["total"]["coverage"]["percentage"]
                        if cov is not None:
                            cov_sum += float(cov)
                            cov_days += 1
                    week_key = f"Uke {week_num}"
                    if valid_days:
                        result.setdefault(week_key, {})[point_id] = total_volume / valid_days
                    if cov_days:
                        cov_result.setdefault(week_key, {})[point_id] = cov_sum / cov_days
            except Exception as e:
                wn, pid = futures[future]
                logger.error("Feil ved henting av ukesdata for uke %s, punkt %s: %s", wn, pid, str(e))
    return result, cov_result


def sum_traffic_data(
    traffic_data_dict: Dict[str, List[Dict]],
    expected_point_ids: Optional[List[str]] = None,
    estimate_missing_points: bool = False,
) -> Tuple[List[float], List[Dict[str, float]], List[bool], List[int]]:
    monthly_sums = [0.0] * 12
    monthly_confidence = [{"lower": 0.0, "upper": 0.0} for _ in range(12)]
    monthly_has_data = [False] * 12
    monthly_points_present = [0] * 12
    expected_points = len(expected_point_ids or [])

    for point_data in (traffic_data_dict or {}).values():
        for entry in point_data or []:
            month = int(entry.get("month") or 0)
            if not (1 <= month <= 12):
                continue
            volume = entry.get("total", {}).get("volume", {}).get("average")
            if volume is None:
                continue
            monthly_has_data[month - 1] = True
            monthly_sums[month - 1] += float(volume)
            monthly_points_present[month - 1] += 1
            ci = entry.get("total", {}).get("volume", {}).get("confidenceInterval") or {}
            lb = ci.get("lowerBound")
            ub = ci.get("upperBound")
            if lb is not None and ub is not None:
                monthly_confidence[month - 1]["lower"] += float(lb)
                monthly_confidence[month - 1]["upper"] += float(ub)

    if estimate_missing_points and expected_points:
        for i in range(12):
            present = monthly_points_present[i]
            if present and present < expected_points:
                monthly_sums[i] = monthly_sums[i] * (expected_points / present)
                monthly_confidence[i]["lower"] = monthly_confidence[i]["lower"] * (expected_points / present)
                monthly_confidence[i]["upper"] = monthly_confidence[i]["upper"] * (expected_points / present)

    return monthly_sums, monthly_confidence, monthly_has_data, monthly_points_present


def detect_monthly_anomalies(df: pd.DataFrame, threshold_pct: float = ANOMALY_THRESHOLD_PCT) -> pd.DataFrame:
    if df is None or df.empty or "Month" not in df.columns:
        return pd.DataFrame()
    years = [c for c in df.columns if str(c).isdigit()]
    if len(years) < 2:
        return pd.DataFrame()

    rows: List[Dict[str, object]] = []
    for year_col in years:
        other_cols = [c for c in years if c != year_col]
        for _, r in df.iterrows():
            month = int(r.get("Month", 0) or 0)
            if not (1 <= month <= 12):
                continue
            actual = pd.to_numeric(r.get(year_col), errors="coerce")
            if pd.isna(actual):
                continue
            expected = pd.to_numeric(pd.Series([r.get(c) for c in other_cols]), errors="coerce").median()
            if pd.isna(expected) or expected == 0:
                continue
            deviation = (float(actual) - float(expected)) / float(expected) * 100.0
            if abs(deviation) >= float(threshold_pct):
                rows.append(
                    {
                        "year": int(year_col),
                        "month": month,
                        "month_name": MONTH_NAMES[month - 1],
                        "actual": float(actual),
                        "expected": float(expected),
                        "deviation_pct": float(deviation),
                    }
                )
    out = pd.DataFrame(rows)
    if out.empty:
        return out
    return out.sort_values(["year", "month"]).reset_index(drop=True)


def sum_weekly_traffic_data(weekly_data_dict: Dict[str, Dict[str, float]]) -> Dict[str, float]:
    return {week: float(sum(vals.values())) for week, vals in weekly_data_dict.items()}


def add_month_names(df: pd.DataFrame) -> pd.DataFrame:
    if "Month" in df.columns:
        df["Month Name"] = [MONTH_NAMES[i - 1] for i in df["Month"]]
        return df[["Month", "Month Name"] + [c for c in df.columns if c not in ["Month", "Month Name"]]]
    return df


def calculate_yearly_total_from_monthly_averages(df: pd.DataFrame, year: int) -> Tuple[float, int, int]:
    year_col = str(year)
    if df is None or df.empty or "Month" not in df.columns or year_col not in df.columns:
        return 0.0, 0, 0

    months = pd.to_numeric(df["Month"], errors="coerce")
    avg_vals = pd.to_numeric(df[year_col], errors="coerce")
    valid = avg_vals.notna() & months.between(1, 12)
    if not valid.any():
        return 0.0, 0, 0

    valid_months = months[valid].astype(int)
    dims = valid_months.apply(lambda m: calendar.monthrange(year, m)[1])
    total = float((avg_vals[valid] * dims).sum())
    return total, int(valid.sum()), int(dims.sum())


def extract_point_monthly_metrics(traffic_by_point: Dict[str, List[Dict]], year: int) -> pd.DataFrame:
    """
    Returnerer per-punkt/per-måned:
      - avg_daily (ÅDT)
      - coverage_pct
      - ci_lower/ci_upper (ÅDT), hvis tilgjengelig
    """
    rows: List[Dict[str, object]] = []
    for point_id, entries in (traffic_by_point or {}).items():
        for entry in entries or []:
            month = int(entry.get("month", 0) or 0)
            if not (1 <= month <= 12):
                continue
            total = entry.get("total", {}) or {}
            volume = (total.get("volume", {}) or {}).get("average")
            cov = (total.get("coverage", {}) or {}).get("percentage")
            ci = (total.get("volume", {}) or {}).get("confidenceInterval") or {}
            rows.append(
                {
                    "point_id": point_id,
                    "point_label": POINT_ID_LABELS.get(point_id, point_id),
                    "year": year,
                    "month": month,
                    "month_name": MONTH_NAMES[month - 1],
                    "avg_daily": float(volume) if volume is not None else np.nan,
                    "coverage_pct": float(cov) if cov is not None else np.nan,
                    "ci_lower": float(ci["lowerBound"]) if ci.get("lowerBound") is not None else np.nan,
                    "ci_upper": float(ci["upperBound"]) if ci.get("upperBound") is not None else np.nan,
                }
            )
    return pd.DataFrame(rows)

def compute_monthly_coverage_summary(
    traffic_by_point: Dict[str, List[Dict]],
    year: int,
    expected_point_ids: List[str],
) -> pd.DataFrame:
    metrics = extract_point_monthly_metrics(traffic_by_point, year)
    expected_points = len(expected_point_ids)

    if metrics.empty:
        return pd.DataFrame(
            {
                "year": [year] * 12,
                "month": list(range(1, 13)),
                "month_name": MONTH_NAMES,
                "points_expected": [expected_points] * 12,
                "points_present": [0] * 12,
                "points_present_pct": [0.0] * 12,
                "mean_coverage_pct": [np.nan] * 12,
                "min_coverage_pct": [np.nan] * 12,
            }
        )

    tmp = metrics[metrics["avg_daily"].notna()].copy()
    grouped = (
        tmp.groupby(["year", "month", "month_name"], as_index=False)
        .agg(
            points_present=("point_id", "nunique"),
            mean_coverage_pct=("coverage_pct", "mean"),
            min_coverage_pct=("coverage_pct", "min"),
        )
        .sort_values("month")
    )
    base = pd.DataFrame({"year": [year] * 12, "month": list(range(1, 13)), "month_name": MONTH_NAMES})
    out = base.merge(grouped, on=["year", "month", "month_name"], how="left")
    out["points_expected"] = expected_points
    out["points_present"] = out["points_present"].fillna(0).astype(int)
    out["points_present_pct"] = np.where(
        expected_points > 0,
        out["points_present"].astype(float) / float(expected_points) * 100.0,
        np.nan,
    )
    return out


def compute_weekly_coverage_summary(
    weekly_data_by_point: Dict[str, Dict[str, float]],
    weekly_cov_by_point: Dict[str, Dict[str, float]],
    expected_point_ids: List[str],
    year: int,
) -> pd.DataFrame:
    expected_points = len(expected_point_ids)
    weeks = sorted(
        {str(w) for w in (weekly_data_by_point or {}).keys()},
        key=lambda s: int("".join([c for c in s if c.isdigit()]) or "0"),
    )
    rows: List[Dict[str, object]] = []
    for week in weeks:
        vols = (weekly_data_by_point or {}).get(week, {}) or {}
        covs = (weekly_cov_by_point or {}).get(week, {}) or {}
        points_present = len(vols)
        cov_values = [float(v) for v in covs.values() if v is not None and not pd.isna(v)]
        rows.append(
            {
                "year": year,
                "week": week,
                "points_expected": expected_points,
                "points_present": points_present,
                "points_present_pct": (points_present / expected_points * 100.0) if expected_points else np.nan,
                "mean_coverage_pct": (sum(cov_values) / len(cov_values)) if cov_values else np.nan,
                "min_coverage_pct": min(cov_values) if cov_values else np.nan,
            }
        )
    return pd.DataFrame(rows)


def coverage_pivot(metrics_df: pd.DataFrame) -> pd.DataFrame:
    if metrics_df is None or metrics_df.empty:
        return pd.DataFrame()
    pivot = metrics_df.pivot_table(
        index="month_name",
        columns="point_label",
        values="coverage_pct",
        aggfunc="mean",
    ).reindex(MONTH_NAMES)
    return pivot


def group_coverage_by_month(metrics_df: pd.DataFrame, point_ids_by_group: Dict[str, List[str]]) -> pd.DataFrame:
    if metrics_df is None or metrics_df.empty:
        return pd.DataFrame()
    group_rows: List[pd.DataFrame] = []
    for group, ids in point_ids_by_group.items():
        labels = {POINT_ID_LABELS.get(pid, pid) for pid in ids}
        sub = metrics_df[metrics_df["point_label"].isin(labels)].copy()
        if sub.empty:
            continue
        grouped = (
            sub.groupby("month_name", as_index=False)
            .agg(coverage_pct=("coverage_pct", "mean"))
            .assign(group=group)
        )
        group_rows.append(grouped)
    if not group_rows:
        return pd.DataFrame()
    out = pd.concat(group_rows, ignore_index=True)
    out["month_name"] = pd.Categorical(out["month_name"], categories=MONTH_NAMES, ordered=True)
    out = out.sort_values(["month_name", "group"])
    return out


def totals_with_uncertainty_from_metrics(metrics_df: pd.DataFrame) -> pd.DataFrame:
    """
    Lager per måned totals (passeringer) og et "min/max"-intervall basert på CI.
    Merk: Summert CI på tvers av punkter er en forenkling (indikativ usikkerhet).
    """
    if metrics_df is None or metrics_df.empty:
        return pd.DataFrame()
    tmp = metrics_df.copy()
    tmp["days_in_month"] = tmp.apply(
        lambda row: calendar.monthrange(int(row["year"]), int(row["month"]))[1], axis=1
    )
    tmp["total"] = tmp["avg_daily"] * tmp["days_in_month"]
    tmp["total_lower"] = tmp["ci_lower"] * tmp["days_in_month"]
    tmp["total_upper"] = tmp["ci_upper"] * tmp["days_in_month"]
    out = (
        tmp.groupby(["month", "month_name"], as_index=False)
        .agg(
            total=("total", "sum"),
            total_lower=("total_lower", "sum"),
            total_upper=("total_upper", "sum"),
            coverage_pct=("coverage_pct", "mean"),
        )
        .sort_values("month")
    )
    return out


def monthly_totals_from_monthly_averages(df: pd.DataFrame, year: int) -> pd.Series:
    year_col = str(year)
    if df is None or df.empty or "Month" not in df.columns or year_col not in df.columns:
        return pd.Series([np.nan] * (len(df) if df is not None else 0))

    months = pd.to_numeric(df["Month"], errors="coerce")
    avg_vals = pd.to_numeric(df[year_col], errors="coerce")
    valid = avg_vals.notna() & months.between(1, 12)
    dims = months.where(valid).apply(
        lambda m: calendar.monthrange(year, int(m))[1] if pd.notna(m) else np.nan
    )
    return (avg_vals * dims).where(valid)


def compute_monthly_totals_table(df: pd.DataFrame, years: List[int]) -> pd.DataFrame:
    base_cols = [c for c in ["Month", "Month Name"] if c in df.columns]
    out = df[base_cols].copy()
    for y in years:
        out[str(y)] = monthly_totals_from_monthly_averages(df, y)
    numeric_columns = out.select_dtypes(include=[np.number]).columns
    out[numeric_columns] = out[numeric_columns].round(0).astype("Int64")
    return out


def aggregate_monthly_totals_by_group(
    traffic_data_by_point: Dict[str, List[Dict]],
    point_ids_by_group: Dict[str, List[str]],
    year: int,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    point_to_group = {pid: group for group, ids in point_ids_by_group.items() for pid in ids}
    groups = list(point_ids_by_group.keys())
    totals = {g: {m: 0.0 for m in range(1, 13)} for g in groups}
    cov_lists: Dict[str, Dict[int, List[float]]] = {g: {m: [] for m in range(1, 13)} for g in groups}

    for pid, entries in traffic_data_by_point.items():
        group = point_to_group.get(pid)
        if not group:
            continue
        for entry in entries or []:
            month = int(entry.get("month", 0) or 0)
            if not (1 <= month <= 12):
                continue
            avg_daily = entry.get("total", {}).get("volume", {}).get("average")
            if avg_daily is None:
                continue
            dim = calendar.monthrange(year, month)[1]
            totals[group][month] += float(avg_daily) * dim
            cov = entry.get("total", {}).get("coverage", {}).get("percentage")
            if cov is not None:
                cov_lists[group][month].append(float(cov))

    totals_df = pd.DataFrame({"Month": list(range(1, 13))})
    for g in groups:
        totals_df[g] = [totals[g][m] for m in range(1, 13)]
    totals_df = add_month_names(totals_df)
    totals_df[groups] = totals_df[groups].round(0).astype("Int64")

    coverage_df = pd.DataFrame({"Month": list(range(1, 13))})
    for g in groups:
        coverage_df[g] = [
            (sum(cov_lists[g][m]) / len(cov_lists[g][m])) if cov_lists[g][m] else np.nan for m in range(1, 13)
        ]
    coverage_df = add_month_names(coverage_df)
    return totals_df, coverage_df


def calculate_growth_rates(df: pd.DataFrame) -> pd.DataFrame:
    growth_df = df.copy()
    year_columns = [col for col in df.columns if col not in ["Month", "Month Name", "Week", "Volume"]]
    if len(year_columns) >= 2:
        for i in range(1, len(year_columns)):
            prev_year = year_columns[i - 1]
            curr_year = year_columns[i]
            growth_col = f"Vekst {prev_year}-{curr_year} (%)"
            curr = pd.to_numeric(df[curr_year], errors="coerce")
            prev_vals = pd.to_numeric(df[prev_year], errors="coerce")
            growth_df[growth_col] = ((curr - prev_vals) / prev_vals * 100).round(1)
    return growth_df


def calculate_seasonal_patterns(df: pd.DataFrame) -> Dict:
    if "Month" not in df.columns:
        return {}
    patterns = {}
    year_columns = [col for col in df.columns if col not in ["Month", "Month Name"]]
    for year in year_columns:
        if year in df.columns:
            yearly = pd.to_numeric(df[year], errors="coerce").to_numpy()
            if len(yearly) == 12 and not np.all(np.isnan(yearly)):
                patterns[year] = {
                    "vinter_snitt": np.nanmean([yearly[11], yearly[0], yearly[1]]),
                    "vår_snitt": np.nanmean(yearly[2:5]),
                    "sommer_snitt": np.nanmean(yearly[5:8]),
                    "høst_snitt": np.nanmean(yearly[8:11]),
                }
    return patterns


def export_to_excel(df: pd.DataFrame) -> bytes:
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is not available")
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Trafikkdata", index=False)
        if len([col for col in df.columns if str(col).isdigit()]) >= 2:
            calculate_growth_rates(df).to_excel(writer, sheet_name="Vekstrater", index=False)
        seasonal = calculate_seasonal_patterns(df)
        if seasonal:
            pd.DataFrame(seasonal).T.to_excel(writer, sheet_name="Sesongmønstre")
    return output.getvalue()

def build_excel_report(
    df: pd.DataFrame,
    point: str,
    comparison_mode: str,
    year_list: List[int],
    year: int,
    point_ids: List[str],
    timeout_s: int,
    use_cache: bool,
    coverage_threshold: float,
    ryfast_include_ramp: Optional[bool] = None,
    estimate_missing_points: bool = False,
) -> bytes:
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is not available")

    include_ramp = (
        bool(st.session_state.get("ryfast_include_ramp", True)) if ryfast_include_ramp is None else bool(ryfast_include_ramp)
    )

    header_font = Font(bold=True)
    title_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
    wrap_top = Alignment(vertical="top", wrap_text=True)
    num_format_int = "#,##0"
    num_format_pct_1 = "0.0"

    def style_header_row(ws, row: int = 1) -> None:
        for cell in ws[row]:
            cell.font = header_font
            cell.fill = header_fill

    wb = Workbook()

    # Summary (first sheet)
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary["A1"] = "Ryfast - rapport"
    ws_summary["A1"].font = title_font
    ws_summary["A2"] = f"Målepunkt: {point}"
    ws_summary["A3"] = f"Analysetype: {comparison_mode}"
    ws_summary["A4"] = f"År: {','.join(map(str, year_list)) if comparison_mode == COMPARE_YEARS else str(year)}"
    ws_summary["A5"] = f"Generert: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_summary["A6"] = f"Min. dekning: {coverage_threshold:.0f}%"
    for r in range(2, 7):
        ws_summary[f"A{r}"].alignment = wrap_top
    ws_summary.column_dimensions["A"].width = 80

    # Metadata
    ws_meta = wb.create_sheet("Metadata")
    ws_meta["A1"] = "Metadata"
    ws_meta["A1"].font = title_font
    meta_rows = [
        ("Generert", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Målepunkt", point),
        ("Analysetype", comparison_mode),
        ("År", ",".join(map(str, year_list)) if comparison_mode == COMPARE_YEARS else str(year)),
        ("Punkt-IDer", ", ".join(point_ids)),
        ("Timeout (s)", str(timeout_s)),
        ("Cache", "Ja" if use_cache else "Nei"),
        ("Min. dekning (%)", f"{coverage_threshold:.0f}"),
        ("Estimer manglende punkt", "Ja" if estimate_missing_points else "Nei"),
    ]
    for i, (k, v) in enumerate(meta_rows, start=3):
        ws_meta[f"A{i}"] = k
        ws_meta[f"B{i}"] = v
        ws_meta[f"A{i}"].font = header_font
        ws_meta[f"B{i}"].alignment = wrap_top
    ws_meta.column_dimensions["A"].width = 20
    ws_meta.column_dimensions["B"].width = 120

    # Summary - yearly totals from df (no extra API calls)
    year_cols_in_df = sorted([int(c) for c in df.columns if str(c).isdigit()])
    if comparison_mode != COMPARE_WEEKS and year_cols_in_df:
        ws_summary["A8"] = "Årsoppsummering (beregnet fra månedsdata)"
        ws_summary["A8"].font = Font(bold=True, size=12)
        ws_summary.append(["År", "Sum (mnd)", "Måneder", "Dager dekket", "Snitt per døgn", "Helårsestimat", "YoY (%)"])
        style_header_row(ws_summary, row=9)

        prev_estimate: Optional[float] = None
        for y in year_cols_in_df:
            total, months_present, days_covered = calculate_yearly_total_from_monthly_averages(df, int(y))
            avg_per_day = (total / days_covered) if days_covered else None
            helar = None
            if avg_per_day is not None and not pd.isna(avg_per_day):
                helar = float(avg_per_day) * days_in_year(int(y))
            yoy = None
            if prev_estimate and helar and prev_estimate != 0:
                yoy = (helar - prev_estimate) / prev_estimate * 100
            ws_summary.append(
                [
                    int(y),
                    float(total) if total is not None else None,
                    int(months_present),
                    int(days_covered),
                    float(avg_per_day) if avg_per_day is not None else None,
                    float(helar) if helar is not None else None,
                    float(yoy) if yoy is not None else None,
                ]
            )
            prev_estimate = helar if helar is not None else prev_estimate

        first_data_row = 10
        last_data_row = 9 + len(year_cols_in_df)
        for r in range(first_data_row, last_data_row + 1):
            ws_summary.cell(row=r, column=2).number_format = num_format_int
            ws_summary.cell(row=r, column=5).number_format = num_format_int
            ws_summary.cell(row=r, column=6).number_format = num_format_int
            ws_summary.cell(row=r, column=7).number_format = num_format_pct_1

        for col, width in [("B", 14), ("C", 10), ("D", 12), ("E", 14), ("F", 14), ("G", 10)]:
            ws_summary.column_dimensions[col].width = width

        try:
            chart = BarChart()
            chart.title = "Helårsestimat (beregnet)"
            chart.y_axis.title = "Passeringer"
            data_ref = Reference(ws_summary, min_col=6, min_row=9, max_row=last_data_row)
            cats = Reference(ws_summary, min_col=1, min_row=10, max_row=last_data_row)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 9
            chart.width = 20
            ws_summary.add_chart(chart, "I8")
        except Exception as exc:
            logger.warning("Excel: klarte ikke legge til årsestimatgraf: %s", exc)

    # Data sheet
    ws_data = wb.create_sheet("Data")
    ws_data.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws_data.append(list(row))
    ws_data.freeze_panes = "A2"
    ws_data.auto_filter.ref = ws_data.dimensions
    style_header_row(ws_data, row=1)

    # Totals sheet (if applicable)
    if comparison_mode != COMPARE_WEEKS:
        years_for_totals = year_list if comparison_mode == COMPARE_YEARS else [year]
        totals_df = compute_monthly_totals_table(df, years_for_totals)
        ws_tot = wb.create_sheet("Totals")
        ws_tot.append(list(totals_df.columns))
        for row in totals_df.itertuples(index=False):
            ws_tot.append(list(row))
        ws_tot.freeze_panes = "A2"
        ws_tot.auto_filter.ref = ws_tot.dimensions
        style_header_row(ws_tot, row=1)
        for col_idx, col_name in enumerate(totals_df.columns, start=1):
            if str(col_name).isdigit():
                for r in range(2, 2 + len(totals_df)):
                    ws_tot.cell(row=r, column=col_idx).number_format = num_format_int

        # Simple chart
        if len(years_for_totals) >= 1:
            chart = LineChart()
            chart.title = "Totale passeringer per måned"
            chart.y_axis.title = "Passeringer"
            chart.x_axis.title = "Måned"
            data_ref = Reference(
                ws_tot,
                min_col=3,
                min_row=1,
                max_col=2 + len(years_for_totals),
                max_row=1 + len(totals_df),
            )
            cats = Reference(ws_tot, min_col=2, min_row=2, max_row=1 + len(totals_df))
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)
            ws_tot.add_chart(chart, "H2")

    # Coverage sheets (on-demand for a selected year)
    if comparison_mode != COMPARE_WEEKS:
        selected_year = max(year_list) if comparison_mode == COMPARE_YEARS else year
        traffic_by_point = fetch_batch_traffic_data(point_ids, int(selected_year), timeout_s, use_cache)
        metrics = extract_point_monthly_metrics(traffic_by_point, int(selected_year))
        if not metrics.empty:
            cov_piv = coverage_pivot(metrics)
            ws_cov = wb.create_sheet("Coverage (per point)")
            ws_cov.append(["Måned"] + list(cov_piv.columns))
            for month in MONTH_NAMES:
                row = [month]
                if month in cov_piv.index:
                    row += [float(x) if pd.notna(x) else None for x in cov_piv.loc[month].tolist()]
                else:
                    row += [None] * len(cov_piv.columns)
                ws_cov.append(row)

            ws_cov.freeze_panes = "B2"
            ws_cov.auto_filter.ref = ws_cov.dimensions
            style_header_row(ws_cov, row=1)
            for r in range(2, 2 + len(MONTH_NAMES)):
                for c in range(2, 2 + len(cov_piv.columns)):
                    ws_cov.cell(row=r, column=c).number_format = num_format_pct_1

            # Conditional formatting: red if below threshold
            if cov_piv.shape[1] > 0:
                start_cell = ws_cov.cell(row=2, column=2).coordinate
                end_cell = ws_cov.cell(row=1 + len(MONTH_NAMES), column=1 + cov_piv.shape[1]).coordinate
                ws_cov.conditional_formatting.add(
                    f"{start_cell}:{end_cell}",
                    CellIsRule(
                        operator="lessThan",
                        formula=[str(float(coverage_threshold))],
                        stopIfTrue=True,
                        font=Font(color="9C0006"),
                    ),
                )

            # Coverage summary (by month)
            try:
                below = metrics[(metrics["coverage_pct"].notna()) & (metrics["coverage_pct"] < float(coverage_threshold))].copy()
                by_month = (
                    metrics.groupby("month_name", as_index=False)
                    .agg(
                        mean_coverage=("coverage_pct", "mean"),
                        min_coverage=("coverage_pct", "min"),
                    )
                    .set_index("month_name")
                    .reindex(MONTH_NAMES)
                )
                below_count = (
                    below.groupby("month_name").size().reindex(MONTH_NAMES).fillna(0).astype(int)
                    if not below.empty
                    else pd.Series([0] * len(MONTH_NAMES), index=MONTH_NAMES)
                )
                worst_point = (
                    metrics.sort_values("coverage_pct", ascending=True)
                    .dropna(subset=["coverage_pct"])
                    .groupby("month_name")["point_label"]
                    .first()
                    .reindex(MONTH_NAMES)
                )

                ws_covsum = wb.create_sheet("Coverage summary")
                ws_covsum.append(
                    ["Måned", "Snitt dekning (%)", "Min dekning (%)", "Under terskel (#)", "Verste punkt", "Terskel (%)"]
                )
                style_header_row(ws_covsum, row=1)
                for month in MONTH_NAMES:
                    ws_covsum.append(
                        [
                            month,
                            float(by_month.loc[month, "mean_coverage"])
                            if month in by_month.index and pd.notna(by_month.loc[month, "mean_coverage"])
                            else None,
                            float(by_month.loc[month, "min_coverage"])
                            if month in by_month.index and pd.notna(by_month.loc[month, "min_coverage"])
                            else None,
                            int(below_count.loc[month]) if month in below_count.index else 0,
                            str(worst_point.loc[month]) if month in worst_point.index and pd.notna(worst_point.loc[month]) else None,
                            float(coverage_threshold),
                        ]
                    )

                ws_covsum.freeze_panes = "A2"
                ws_covsum.auto_filter.ref = ws_covsum.dimensions
                ws_covsum.column_dimensions["A"].width = 14
                ws_covsum.column_dimensions["B"].width = 18
                ws_covsum.column_dimensions["C"].width = 16
                ws_covsum.column_dimensions["D"].width = 16
                ws_covsum.column_dimensions["E"].width = 30
                ws_covsum.column_dimensions["F"].width = 12
                for r in range(2, 2 + len(MONTH_NAMES)):
                    ws_covsum.cell(row=r, column=2).number_format = num_format_pct_1
                    ws_covsum.cell(row=r, column=3).number_format = num_format_pct_1
                    ws_covsum.cell(row=r, column=6).number_format = num_format_pct_1

                chart = LineChart()
                chart.title = f"Dekning per måned ({selected_year})"
                chart.y_axis.title = "Dekning (%)"
                chart.x_axis.title = "Måned"
                data_ref = Reference(ws_covsum, min_col=2, min_row=1, max_col=6, max_row=1 + len(MONTH_NAMES))
                cats = Reference(ws_covsum, min_col=1, min_row=2, max_row=1 + len(MONTH_NAMES))
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats)
                chart.height = 9
                chart.width = 20
                ws_covsum.add_chart(chart, "H2")

                if not below.empty:
                    ws_low = wb.create_sheet("Low coverage")
                    ws_low.append(["År", "Måned", "Målepunkt", "Dekning (%)"])
                    style_header_row(ws_low, row=1)
                    below = below.sort_values(["coverage_pct", "month", "point_label"], ascending=[True, True, True])
                    for r in below.itertuples(index=False):
                        ws_low.append(
                            [
                                int(selected_year),
                                str(r.month_name),
                                str(r.point_label),
                                float(r.coverage_pct) if pd.notna(r.coverage_pct) else None,
                            ]
                        )
                    ws_low.freeze_panes = "A2"
                    ws_low.auto_filter.ref = ws_low.dimensions
                    ws_low.column_dimensions["A"].width = 8
                    ws_low.column_dimensions["B"].width = 14
                    ws_low.column_dimensions["C"].width = 30
                    ws_low.column_dimensions["D"].width = 12
                    for rr in range(2, 2 + len(below)):
                        ws_low.cell(row=rr, column=4).number_format = num_format_pct_1
            except Exception as exc:
                logger.warning("Excel: lav-dekning-ark feilet: %s", exc)

            if point == "Ryfast (sum tunneler)":
                point_ids_by_group = {
                    "Ryfylketunnelen": TRAFFIC_POINTS["Ryfylketunnelen"]["ids"],
                    "Hundvågtunnelen": (
                        TRAFFIC_POINTS["Hundvågtunnelen"]["ids"] if include_ramp else HUNDVAG_TUNNEL_IDS_UTEN_PÅRAMPE
                    ),
                }

                grouped = group_coverage_by_month(metrics, point_ids_by_group)
                if not grouped.empty:
                    ws_g = wb.create_sheet("Coverage (tunnel)")
                    ws_g.append(["Måned", "Tunnel", "Dekning (%)"])
                    for row in grouped.itertuples(index=False):
                        ws_g.append([row.month_name, row.group, float(row.coverage_pct) if pd.notna(row.coverage_pct) else None])
                    ws_g.freeze_panes = "A2"
                    ws_g.auto_filter.ref = ws_g.dimensions
                    style_header_row(ws_g, row=1)
                    for rr in range(2, 2 + len(grouped)):
                        ws_g.cell(row=rr, column=3).number_format = num_format_pct_1

                try:
                    totals_by_group_df, coverage_by_group_df = aggregate_monthly_totals_by_group(
                        traffic_by_point, point_ids_by_group, int(selected_year)
                    )
                    if totals_by_group_df is not None and not totals_by_group_df.empty:
                        ws_ttot = wb.create_sheet("Tunnel totals")
                        ws_ttot.append(list(totals_by_group_df.columns))
                        for row in totals_by_group_df.itertuples(index=False):
                            ws_ttot.append(list(row))
                        ws_ttot.freeze_panes = "A2"
                        ws_ttot.auto_filter.ref = ws_ttot.dimensions
                        style_header_row(ws_ttot, row=1)
                        for col_idx, col_name in enumerate(totals_by_group_df.columns, start=1):
                            if col_name in list(point_ids_by_group.keys()):
                                for rr in range(2, 2 + len(totals_by_group_df)):
                                    ws_ttot.cell(row=rr, column=col_idx).number_format = num_format_int

                        chart = LineChart()
                        chart.title = f"Tunnel-fordeling per måned ({selected_year})"
                        chart.y_axis.title = "Passeringer"
                        chart.x_axis.title = "Måned"
                        data_ref = Reference(
                            ws_ttot,
                            min_col=3,
                            min_row=1,
                            max_col=2 + len(point_ids_by_group),
                            max_row=1 + len(totals_by_group_df),
                        )
                        cats = Reference(ws_ttot, min_col=2, min_row=2, max_row=1 + len(totals_by_group_df))
                        chart.add_data(data_ref, titles_from_data=True)
                        chart.set_categories(cats)
                        ws_ttot.add_chart(chart, "H2")

                    if coverage_by_group_df is not None and not coverage_by_group_df.empty:
                        ws_tcov = wb.create_sheet("Tunnel coverage")
                        ws_tcov.append(list(coverage_by_group_df.columns))
                        for row in coverage_by_group_df.itertuples(index=False):
                            ws_tcov.append(list(row))
                        ws_tcov.freeze_panes = "A2"
                        ws_tcov.auto_filter.ref = ws_tcov.dimensions
                        style_header_row(ws_tcov, row=1)
                        for col_idx, col_name in enumerate(coverage_by_group_df.columns, start=1):
                            if col_name in list(point_ids_by_group.keys()):
                                for rr in range(2, 2 + len(coverage_by_group_df)):
                                    ws_tcov.cell(row=rr, column=col_idx).number_format = num_format_pct_1
                except Exception as exc:
                    logger.warning("Excel: tunnel-ark feilet: %s", exc)

            # Data quality warnings
            try:
                warnings: List[Tuple[str, str, Optional[str]]] = []

                cov_month = compute_monthly_coverage_summary(traffic_by_point, int(selected_year), point_ids)
                missing_points = cov_month[(cov_month["points_present"] < cov_month["points_expected"]) & (cov_month["points_present"] > 0)]
                for r in missing_points.itertuples(index=False):
                    warnings.append(
                        (
                            "Manglende målepunkter",
                            f"{int(r.year)} {str(r.month_name)}",
                            f"{int(r.points_present)}/{int(r.points_expected)} punkter med data",
                        )
                    )
                none_points = cov_month[cov_month["points_present"] == 0]
                for r in none_points.itertuples(index=False):
                    warnings.append(("Ingen data", f"{int(r.year)} {str(r.month_name)}", None))

                anomalies = detect_monthly_anomalies(df, threshold_pct=float(ANOMALY_THRESHOLD_PCT))
                for r in anomalies.itertuples(index=False):
                    warnings.append(
                        (
                            "Anomali",
                            f"{int(r.year)} {str(r.month_name)}",
                            f"Avvik {float(r.deviation_pct):+.1f}% (faktisk {int(round(float(r.actual))):,} / forventet {int(round(float(r.expected))):,})".replace(
                                ",", " "
                            ),
                        )
                    )

                if warnings:
                    ws_warn = wb.create_sheet("Warnings")
                    ws_warn.append(["Type", "Periode", "Detalj"])
                    style_header_row(ws_warn, row=1)
                    for t, period, detail in warnings:
                        ws_warn.append([t, period, detail])
                    ws_warn.freeze_panes = "A2"
                    ws_warn.auto_filter.ref = ws_warn.dimensions
                    ws_warn.column_dimensions["A"].width = 22
                    ws_warn.column_dimensions["B"].width = 18
                    ws_warn.column_dimensions["C"].width = 70

                    ws_summary["A7"] = f"Varsler: {len(warnings)} (se fanen 'Warnings')"
                    ws_summary["A7"].alignment = wrap_top
            except Exception as exc:
                logger.warning("Excel: varsels-ark feilet: %s", exc)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _pdf_embed_figure(pdf: "FPDF", fig: "go.Figure") -> None:
    """Render a Plotly figure as PNG and embed it in the PDF, cleaning up the temp file."""
    img_path: Optional[str] = None
    try:
        img_bytes = fig.to_image(format="png", scale=2)
        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
            tmp.write(img_bytes)
            img_path = tmp.name
        pdf.ln(2)
        pdf.image(img_path, w=190)
    except Exception as exc:
        logger.warning("PDF: klarte ikke legge til figur: %s", exc)
    finally:
        if img_path:
            try:
                os.remove(img_path)
            except OSError:
                pass


def build_pdf_report(
    df: pd.DataFrame,
    point: str,
    comparison_mode: str,
    year_list: List[int],
    year: int,
    point_ids: List[str],
    timeout_s: int,
    use_cache: bool,
    coverage_threshold: float,
    ryfast_include_ramp: Optional[bool] = None,
    estimate_missing_points: bool = False,
) -> bytes:
    def pdf_safe_text(text: str) -> str:
        replacements = {
            "\u2013": "-",  # en dash
            "\u2014": "-",  # em dash
            "\u2212": "-",  # minus sign
            "\u00a0": " ",  # nbsp
            "\u2018": "'",  # left single quote
            "\u2019": "'",  # right single quote
            "\u201c": '"',  # left double quote
            "\u201d": '"',  # right double quote
            "\u2026": "...",  # ellipsis
        }
        for src, dst in replacements.items():
            text = text.replace(src, dst)
        return text.encode("latin-1", errors="replace").decode("latin-1")

    include_ramp = (
        bool(st.session_state.get("ryfast_include_ramp", True)) if ryfast_include_ramp is None else bool(ryfast_include_ramp)
    )

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, pdf_safe_text("Ryfast - rapport"), ln=True)
    pdf.set_font("Helvetica", "", 11)
    pdf.multi_cell(
        0,
        6,
        pdf_safe_text(
            f"Generert: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
            f"Målepunkt: {point}\n"
            f"Analysetype: {comparison_mode}\n"
            f"År: {','.join(map(str, year_list)) if comparison_mode == COMPARE_YEARS else str(year)}\n"
            f"Min. dekning: {coverage_threshold:.0f}%\n"
            f"Estimer manglende punkt: {'Ja' if estimate_missing_points else 'Nei'}\n"
            f"Punkt-IDer: {', '.join(point_ids)}"
        ),
    )

    # Year summary (derived from df; no extra API calls)
    year_cols_in_df = sorted([int(c) for c in df.columns if str(c).isdigit()])
    if comparison_mode != COMPARE_WEEKS and year_cols_in_df:
        pdf.ln(2)
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 8, pdf_safe_text("Årsoppsummering (beregnet fra månedsdata)"), ln=True)
        pdf.set_font("Helvetica", "", 10)
        estimates: List[Tuple[int, float]] = []
        prev_estimate: Optional[float] = None
        for y in year_cols_in_df:
            total, months_present, days_covered = calculate_yearly_total_from_monthly_averages(df, int(y))
            avg_per_day = (total / days_covered) if days_covered else None
            estimate = None
            if avg_per_day is not None and not pd.isna(avg_per_day):
                estimate = float(avg_per_day) * days_in_year(int(y))
                estimates.append((int(y), estimate))
            yoy_txt = ""
            if prev_estimate and estimate and prev_estimate != 0:
                yoy = (estimate - prev_estimate) / prev_estimate * 100
                yoy_txt = f"  YoY {yoy:+.1f}%"
            prev_estimate = estimate if estimate is not None else prev_estimate
            pdf.cell(
                0,
                6,
                pdf_safe_text(
                    f"{int(y)}: sum {int(round(total)):,} (mnd={int(months_present)}, dager={int(days_covered)})"
                    + (f"  helår {int(round(estimate)):,}" if estimate is not None else "")
                    + yoy_txt
                ),
                ln=True,
            )

        if len(estimates) >= 1:
            try:
                fig = go.Figure()
                fig.add_trace(go.Bar(x=[str(y) for y, _ in estimates], y=[v for _, v in estimates], name="Helårsestimat"))
                fig.update_layout(
                    title="Helårsestimat (beregnet)",
                    yaxis_title="Passeringer",
                    xaxis_title="År",
                    template="plotly_white",
                    height=320,
                    margin=dict(l=10, r=10, t=40, b=10),
                )
                _pdf_embed_figure(pdf, fig)
            except Exception as exc:
                logger.warning("PDF: helårsestimat-graf feilet: %s", exc)

    if comparison_mode != COMPARE_WEEKS:
        selected_year = max(year_list) if comparison_mode == COMPARE_YEARS else year
        traffic_by_point = fetch_batch_traffic_data(point_ids, int(selected_year), timeout_s, use_cache)
        metrics = extract_point_monthly_metrics(traffic_by_point, int(selected_year))
        totals_ci = totals_with_uncertainty_from_metrics(metrics)
        if not totals_ci.empty:
            try:
                totals_ci_sorted = totals_ci.sort_values("month")
                fig = go.Figure()
                fig.add_trace(go.Bar(x=totals_ci_sorted["month_name"], y=totals_ci_sorted["total"], name="Totalt"))
                fig.add_trace(
                    go.Scatter(
                        x=totals_ci_sorted["month_name"],
                        y=totals_ci_sorted["total_lower"],
                        mode="lines",
                        line=dict(width=0),
                        showlegend=False,
                        hoverinfo="skip",
                    )
                )
                fig.add_trace(
                    go.Scatter(
                        x=totals_ci_sorted["month_name"],
                        y=totals_ci_sorted["total_upper"],
                        mode="lines",
                        line=dict(width=0),
                        fill="tonexty",
                        fillcolor="rgba(31,119,180,0.18)",
                        name="Usikkerhet (indikativ)",
                    )
                )
                fig.update_layout(
                    title=f"Totale passeringer per måned ({selected_year})",
                    yaxis_title="Passeringer",
                    xaxis_title="Måned",
                    template="plotly_white",
                    height=360,
                    margin=dict(l=10, r=10, t=40, b=10),
                )
                _pdf_embed_figure(pdf, fig)
            except Exception as exc:
                logger.warning("PDF: totale passeringer-graf feilet: %s", exc)
        if not totals_ci.empty:
            pdf.ln(2)
            pdf.set_font("Helvetica", "B", 12)
            pdf.cell(
                0,
                8,
                pdf_safe_text(f"Totale passeringer (indikativ usikkerhet) - {selected_year}"),
                ln=True,
            )
            pdf.set_font("Helvetica", "", 10)
            for _, r in totals_ci.iterrows():
                if pd.isna(r["total"]):
                    continue
                pdf.cell(
                    0,
                    6,
                    pdf_safe_text(
                        f"{r['month_name']}: {int(round(r['total'])):,}  "
                        f"[{int(round(r['total_lower'])):,} – {int(round(r['total_upper'])):,}]  "
                        f"dekning {r['coverage_pct']:.1f}%"
                    ),
                    ln=True,
                )

        # Coverage summary (uses already-fetched metrics)
        if metrics is not None and not metrics.empty and metrics["coverage_pct"].notna().any():
            pdf.ln(2)
            pdf.set_font("Helvetica", "B", 12)
            pdf.cell(0, 8, pdf_safe_text(f"Dekning og datakvalitet - {selected_year}"), ln=True)
            pdf.set_font("Helvetica", "", 10)
            below = metrics[(metrics["coverage_pct"].notna()) & (metrics["coverage_pct"] < float(coverage_threshold))].copy()
            pdf.cell(0, 6, pdf_safe_text(f"Snitt dekning: {metrics['coverage_pct'].mean():.1f}%"), ln=True)
            pdf.cell(0, 6, pdf_safe_text(f"Observasjoner under terskel ({coverage_threshold:.0f}%): {len(below)}"), ln=True)

            if not below.empty:
                worst = below.sort_values("coverage_pct", ascending=True).head(8)
                pdf.ln(1)
                pdf.cell(0, 6, pdf_safe_text("Lavest dekning (utvalg):"), ln=True)
                for r in worst.itertuples(index=False):
                    pdf.cell(0, 6, pdf_safe_text(f"- {r.month_name}: {r.point_label} {float(r.coverage_pct):.1f}%"), ln=True)

            if not totals_ci.empty and totals_ci["coverage_pct"].notna().any():
                try:
                    fig = go.Figure()
                    fig.add_trace(
                        go.Scatter(
                            x=totals_ci["month_name"],
                            y=totals_ci["coverage_pct"],
                            mode="lines+markers",
                            name="Dekning (%)",
                        )
                    )
                    fig.add_trace(
                        go.Scatter(
                            x=totals_ci["month_name"],
                            y=[float(coverage_threshold)] * len(totals_ci),
                            mode="lines",
                            name="Terskel",
                            line=dict(dash="dash"),
                        )
                    )
                    fig.update_layout(
                        title=f"Dekning per måned ({selected_year})",
                        yaxis_title="Dekning (%)",
                        xaxis_title="Måned",
                        template="plotly_white",
                        height=320,
                        margin=dict(l=10, r=10, t=40, b=10),
                    )
                    _pdf_embed_figure(pdf, fig)
                except Exception as exc:
                    logger.warning("PDF: dekning-graf feilet: %s", exc)

        # Ryfast tunnel breakdown (optional)
        if point == "Ryfast (sum tunneler)":
            try:
                point_ids_by_group = {
                    "Ryfylketunnelen": TRAFFIC_POINTS["Ryfylketunnelen"]["ids"],
                    "Hundvågtunnelen": (
                        TRAFFIC_POINTS["Hundvågtunnelen"]["ids"] if include_ramp else HUNDVAG_TUNNEL_IDS_UTEN_PÅRAMPE
                    ),
                }
                totals_by_group_df, cov_by_group_df = aggregate_monthly_totals_by_group(
                    traffic_by_point, point_ids_by_group, int(selected_year)
                )
                if totals_by_group_df is not None and not totals_by_group_df.empty:
                    try:
                        fig = px.bar(
                            totals_by_group_df.melt(
                                id_vars=["Month", "Month Name"],
                                value_vars=list(point_ids_by_group.keys()),
                                var_name="Tunnel",
                                value_name="Passeringer",
                            ),
                            x="Month Name",
                            y="Passeringer",
                            color="Tunnel",
                            barmode="stack",
                            title=f"Samlet trafikk per måned ({selected_year}) - tunnel-fordeling",
                        )
                        fig.update_yaxes(tickformat=",")
                        fig.update_layout(template="plotly_white", height=340, margin=dict(l=10, r=10, t=40, b=10))
                        _pdf_embed_figure(pdf, fig)
                    except Exception as exc:
                        logger.warning("PDF: tunnelfordeling-graf feilet: %s", exc)
                if cov_by_group_df is not None and not cov_by_group_df.empty:
                    pdf.ln(2)
                    pdf.set_font("Helvetica", "B", 12)
                    pdf.cell(0, 8, pdf_safe_text(f"Dekning per tunnel (snitt) - {selected_year}"), ln=True)
                    pdf.set_font("Helvetica", "", 10)
                    for r in cov_by_group_df.itertuples(index=False):
                        try:
                            month_name = r._asdict().get("Month Name", "")
                            ryf = getattr(r, "Ryfylketunnelen", np.nan)
                            hund = getattr(r, "Hundvågtunnelen", np.nan)
                            if pd.isna(ryf) and pd.isna(hund):
                                continue
                            pdf.cell(0, 6, pdf_safe_text(f"{month_name}: Ryfylke {float(ryf):.1f}%  Hundvåg {float(hund):.1f}%"), ln=True)
                        except Exception:
                            continue
            except Exception:
                pass

        # Data quality warnings
        try:
            warnings: List[str] = []
            cov_month = compute_monthly_coverage_summary(traffic_by_point, int(selected_year), point_ids)
            missing_months = cov_month[(cov_month["points_present"] < cov_month["points_expected"]) & (cov_month["points_present"] > 0)][
                "month_name"
            ].tolist()
            if missing_months:
                warnings.append("Manglende målepunkter: " + ", ".join([m for m in MONTH_NAMES if m in set(missing_months)]))

            none_months = cov_month[cov_month["points_present"] == 0]["month_name"].tolist()
            if none_months:
                warnings.append("Ingen data: " + ", ".join([m for m in MONTH_NAMES if m in set(none_months)]))

            anomalies = detect_monthly_anomalies(df, threshold_pct=float(ANOMALY_THRESHOLD_PCT))
            if not anomalies.empty:
                for r in anomalies.itertuples(index=False):
                    warnings.append(f"Anomali {int(r.year)} {str(r.month_name)}: {float(r.deviation_pct):+.1f}%")

            if warnings:
                pdf.ln(2)
                pdf.set_font("Helvetica", "B", 12)
                pdf.cell(0, 8, pdf_safe_text("Data Quality Warnings"), ln=True)
                pdf.set_font("Helvetica", "", 10)
                for w in warnings[:15]:
                    pdf.multi_cell(0, 5, pdf_safe_text(f"- {w}"))
        except Exception as exc:
            logger.warning("PDF: feil ved generering av varselseksjon: %s", exc)

    output = pdf.output(dest="S")
    return bytes(output) if isinstance(output, (bytes, bytearray)) else output.encode("latin-1")


def create_export_section(df: pd.DataFrame, point: str, coverage_summary: Optional[pd.DataFrame] = None):
    st.subheader("📊 Eksporter data")
    col1, col2, col3 = st.columns(3)

    with col1:
        csv_data = df.to_csv(index=False, sep=";", encoding="utf-8")
        st.download_button(
            label="📁 Last ned CSV",
            data=csv_data,
            file_name=f"{point}_trafikkdata_{datetime.now().strftime('%Y%m%d')}.csv",
            mime="text/csv",
        )

    with col2:
        json_data = df.to_json(orient="records", indent=2)
        st.download_button(
            label="🔗 Last ned JSON",
            data=json_data,
            file_name=f"{point}_trafikkdata_{datetime.now().strftime('%Y%m%d')}.json",
            mime="application/json",
        )

    with col3:
        if OPENPYXL_AVAILABLE:
            st.download_button(
                label="📄 Last ned Excel",
                data=export_to_excel(df),
                file_name=f"{point}_trafikkdata_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        else:
            st.caption("Excel-export krever `openpyxl`.")

    if coverage_summary is not None and not coverage_summary.empty:
        st.markdown("#### 🛡️ Eksporter dekning")
        c1, c2 = st.columns(2)
        with c1:
            cov_csv = coverage_summary.to_csv(index=False, sep=";", encoding="utf-8")
            st.download_button(
                label="📁 Last ned dekning (CSV)",
                data=cov_csv,
                file_name=f"{point}_dekning_{datetime.now().strftime('%Y%m%d')}.csv",
                mime="text/csv",
            )
        with c2:
            cov_json = coverage_summary.to_json(orient="records", indent=2)
            st.download_button(
                label="🔗 Last ned dekning (JSON)",
                data=cov_json,
                file_name=f"{point}_dekning_{datetime.now().strftime('%Y%m%d')}.json",
                mime="application/json",
            )

    st.markdown("### 📦 Ferdig rapport (anbefalt)")
    st.caption("Inkluderer metadata, data, dekning og (der mulig) grafer i Excel/PDF.")
    c1, c2 = st.columns(2)
    with c1:
        if st.button("📊 Lag Excel-rapport", type="secondary"):
            result = st.session_state.get("last_result") or {}
            report_bytes = build_excel_report(
                df=df,
                point=point,
                comparison_mode=result.get("comparison_mode", ""),
                year_list=result.get("year_list", []),
                year=int(result.get("year", 0) or 0),
                point_ids=result.get("point_ids", []),
                timeout_s=int(result.get("timeout_s", 60)),
                use_cache=bool(result.get("use_cache", True)),
                coverage_threshold=float(result.get("coverage_threshold", 90)),
                ryfast_include_ramp=bool(st.session_state.get("ryfast_include_ramp", True)),
                estimate_missing_points=bool(result.get("estimate_missing_points", False)),
            )
            st.download_button(
                "⬇️ Last ned Excel-rapport",
                data=report_bytes,
                file_name=f"ryfast_rapport_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
    with c2:
        if st.button("📄 Lag PDF-rapport", type="secondary"):
            result = st.session_state.get("last_result") or {}
            report_bytes = build_pdf_report(
                df=df,
                point=point,
                comparison_mode=result.get("comparison_mode", ""),
                year_list=result.get("year_list", []),
                year=int(result.get("year", 0) or 0),
                point_ids=result.get("point_ids", []),
                timeout_s=int(result.get("timeout_s", 60)),
                use_cache=bool(result.get("use_cache", True)),
                coverage_threshold=float(result.get("coverage_threshold", 90)),
                ryfast_include_ramp=bool(st.session_state.get("ryfast_include_ramp", True)),
                estimate_missing_points=bool(result.get("estimate_missing_points", False)),
            )
            st.download_button(
                "⬇️ Last ned PDF-rapport",
                data=report_bytes,
                file_name=f"ryfast_rapport_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
                mime="application/pdf",
            )


def create_advanced_visualization(df: pd.DataFrame, point: str, chart_type: str) -> go.Figure:
    template = "plotly_white"

    if chart_type == "line_with_confidence":
        fig = go.Figure()
        year_columns = [col for col in df.columns if col not in ["Month", "Month Name", "Week", "Volume"]]
        colors = px.colors.qualitative.Set1
        x_vals = df["Month Name"].tolist() if "Month Name" in df.columns else list(df.index)
        for i, year in enumerate(year_columns):
            y = pd.to_numeric(df[year], errors="coerce")
            fig.add_trace(
                go.Scatter(
                    x=x_vals,
                    y=y,
                    mode="lines+markers",
                    name=str(year),
                    line=dict(color=colors[i % len(colors)], width=2.5),
                    marker=dict(size=7, symbol="circle"),
                    hovertemplate="%{x}: <b>%{y:,.0f}</b> ÅDT<extra>%{fullData.name}</extra>",
                )
            )
        fig.update_layout(
            title=dict(text=f"Trafikkutvikling for {point}", font=dict(size=15)),
            xaxis_title="Måned",
            yaxis=dict(title="Gjennomsnittlig døgntrafikk (ÅDT)", tickformat=","),
            hovermode="x unified",
            legend=dict(orientation="h", yanchor="bottom", y=1.02),
            template=template,
            height=520,
            margin=dict(t=60, b=40),
        )
        return fig

    if chart_type == "heatmap":
        year_columns = [col for col in df.columns if col not in ["Month", "Month Name", "Week", "Volume"]]
        if len(year_columns) > 1 and "Month Name" in df.columns:
            table = df[["Month Name"] + year_columns].copy()
            z = table[year_columns].apply(pd.to_numeric, errors="coerce").to_numpy(dtype=float).T
            x = table["Month Name"].tolist()
            y = [str(c) for c in year_columns]

            fig = go.Figure(
                data=go.Heatmap(
                    z=z,
                    x=x,
                    y=y,
                    colorscale="RdYlBu_r",
                    colorbar=dict(title="ÅDT"),
                    hoverongaps=False,
                    text=[[f"{v:,.0f}" if not np.isnan(v) else "" for v in row] for row in z],
                    texttemplate="%{text}",
                    textfont=dict(size=10),
                    hovertemplate="%{y} – %{x}: <b>%{z:,.0f}</b> ÅDT<extra></extra>",
                )
            )
            fig.update_layout(
                title=dict(text=f"Sesongmønster for {point}", font=dict(size=15)),
                xaxis_title="Måned",
                yaxis_title="År",
                template=template,
                height=max(320, 120 * len(y) + 80),
            )
            return fig
        return create_advanced_visualization(df, point, "line")

    if chart_type == "box":
        year_columns = [col for col in df.columns if col not in ["Month", "Month Name", "Week", "Volume"]]
        if not year_columns:
            return create_advanced_visualization(df, point, "line")

        fig = go.Figure()
        colors = px.colors.qualitative.Set1
        for i, year in enumerate(year_columns):
            yvals = pd.to_numeric(df[year], errors="coerce").to_numpy(dtype=float)
            yvals = [v for v in yvals.tolist() if not np.isnan(v)]
            if not yvals:
                continue
            fig.add_trace(
                go.Box(
                    y=yvals,
                    name=str(year),
                    boxpoints="all",
                    jitter=0.3,
                    marker=dict(color=colors[i % len(colors)], size=5, opacity=0.55),
                    line=dict(color=colors[i % len(colors)], width=2),
                    hovertemplate="%{y:,.0f} ÅDT<extra>%{fullData.name}</extra>",
                )
            )
        if not fig.data:
            return create_advanced_visualization(df, point, "line")
        fig.update_layout(
            title=dict(text=f"Trafikkfordeling for {point}", font=dict(size=15)),
            yaxis=dict(title="Gjennomsnittlig døgntrafikk (ÅDT)", tickformat=","),
            xaxis_title="År",
            template=template,
            height=520,
            margin=dict(t=60, b=40),
        )
        return fig

    # Default line chart
    if "Month Name" in df.columns:
        df_melted = df.melt(id_vars=["Month", "Month Name"], var_name="År", value_name="Trafikk")
        x_col = "Month Name"
    elif "Week" in df.columns:
        df_melted = df.melt(id_vars=["Week"], var_name="År", value_name="Trafikk")
        x_col = "Week"
    else:
        df_melted = df.melt(var_name="År", value_name="Trafikk")
        x_col = df_melted.index

    fig = px.line(
        df_melted,
        x=x_col,
        y="Trafikk",
        color="År",
        markers=True,
        title=f"Trafikkutvikling for {point}",
        labels={"Trafikk": "Gjennomsnittlig døgntrafikk (ÅDT)", x_col: "Periode"},
    )
    fig.update_traces(hovertemplate="%{x}: <b>%{y:,.0f}</b> ÅDT<extra>%{fullData.name}</extra>")
    fig.update_layout(
        template=template,
        height=520,
        hovermode="x unified",
        yaxis=dict(tickformat=","),
        legend=dict(orientation="h", yanchor="bottom", y=1.02),
        margin=dict(t=60, b=40),
    )
    return fig


def _year_columns(df: pd.DataFrame) -> List[str]:
    return [col for col in df.columns if col not in ["Month", "Month Name", "Week", "Volume"]]


def _long_year_df(df: pd.DataFrame) -> pd.DataFrame:
    years = _year_columns(df)
    if "Month Name" in df.columns:
        id_vars = ["Month", "Month Name"]
    elif "Week" in df.columns:
        id_vars = ["Week"]
    else:
        id_vars = []
    melted = df.melt(id_vars=id_vars, value_vars=years, var_name="År", value_name="Trafikk")
    melted["Trafikk"] = pd.to_numeric(melted["Trafikk"], errors="coerce")
    return melted.dropna(subset=["Trafikk"])


def _render_altair_chart(df: pd.DataFrame, point: str, chart_type: str):
    years = _year_columns(df)
    if not years:
        st.warning("Fant ingen årskolonner i datasettet. Viser linjediagram.")
        st.plotly_chart(create_advanced_visualization(df, point, "line"), use_container_width=True, key="fallback_line_no_years")
        return

    long_df = _long_year_df(df)
    if long_df.empty:
        st.warning("Fant ingen plottbare data. Viser linjediagram.")
        st.plotly_chart(create_advanced_visualization(df, point, "line"), use_container_width=True, key="fallback_line_empty")
        return

    if chart_type == "heatmap":
        if "Month Name" not in long_df.columns or len(years) < 2:
            st.info("Varmekart krever minst 2 år og månedsdata.")
            st.plotly_chart(create_advanced_visualization(df, point, "line"), use_container_width=True, key="fallback_line_no_heatmap")
            return

        month_sort = MONTH_NAMES
        rect = (
            alt.Chart(long_df)
            .mark_rect()
            .encode(
                x=alt.X("Month Name:N", sort=month_sort, title="Måned"),
                y=alt.Y("År:N", sort=sorted([str(y) for y in years]), title="År"),
                color=alt.Color("Trafikk:Q", scale=alt.Scale(scheme="redyellowblue", reverse=True), title="ÅDT"),
                tooltip=[
                    alt.Tooltip("År:N"),
                    alt.Tooltip("Month Name:N", title="Måned"),
                    alt.Tooltip("Trafikk:Q", format=",.0f", title="ÅDT"),
                ],
            )
        )
        text = (
            alt.Chart(long_df)
            .mark_text(fontSize=9)
            .encode(
                x=alt.X("Month Name:N", sort=month_sort),
                y=alt.Y("År:N", sort=sorted([str(y) for y in years])),
                text=alt.Text("Trafikk:Q", format=",.0f"),
                color=alt.condition(
                    alt.datum.Trafikk > long_df["Trafikk"].median(),
                    alt.value("white"),
                    alt.value("#333"),
                ),
            )
        )
        chart = (rect + text).properties(title=f"Sesongmønster for {point}", height=420)
        st.altair_chart(chart, use_container_width=True)
        return

    if chart_type == "box":
        chart = (
            alt.Chart(long_df)
            .mark_boxplot(extent="min-max")
            .encode(
                x=alt.X("År:N", sort=sorted([str(y) for y in years]), title="År"),
                y=alt.Y("Trafikk:Q", title="Gjennomsnittlig døgntrafikk"),
                color=alt.Color("År:N", legend=None),
                tooltip=[alt.Tooltip("År:N"), alt.Tooltip("Trafikk:Q", format=",.0f")],
            )
            .properties(title=f"Trafikkfordeling for {point}", height=420)
        )
        st.altair_chart(chart, use_container_width=True)
        return

    if chart_type == "line_with_confidence":
        if "Month Name" in long_df.columns:
            x = alt.X("Month Name:N", sort=MONTH_NAMES, title="Måned")
        elif "Week" in long_df.columns:
            x = alt.X("Week:N", title="Uke")
        else:
            x = alt.X("index:O", title="Periode")

        chart = (
            alt.Chart(long_df)
            .mark_line(point=alt.OverlayMarkDef(size=60))
            .encode(
                x=x,
                y=alt.Y("Trafikk:Q", title="Gjennomsnittlig døgntrafikk (ÅDT)", axis=alt.Axis(format=",d")),
                color=alt.Color("År:N", sort=sorted([str(y) for y in years]), title="År"),
                tooltip=[
                    alt.Tooltip("År:N", title="År"),
                    alt.Tooltip("Month Name:N", title="Måned") if "Month Name" in long_df.columns else alt.Tooltip("Week:N", title="Uke"),
                    alt.Tooltip("Trafikk:Q", format=",.0f", title="ÅDT"),
                ],
            )
            .properties(title=f"Trafikkutvikling for {point}", height=460)
        )
        st.altair_chart(chart, use_container_width=True)
        return

    st.plotly_chart(create_advanced_visualization(df, point, "line"), use_container_width=True, key="fallback_line_default")


def create_comparison_dashboard(df: pd.DataFrame, point: str):
    col1, col2 = st.columns(2)
    with col1:
        is_weekly = "Volume" in df.columns
        year_columns = _year_columns(df)
        can_heatmap = (not is_weekly) and ("Month Name" in df.columns) and (len(year_columns) > 1)

        if is_weekly:
            chart_options = ["line"]
        else:
            chart_options = ["line", "box", "line_with_confidence"]
            if can_heatmap:
                chart_options.insert(1, "heatmap")

        chart_type = st.selectbox(
            "Velg diagramtype",
            chart_options,
            key="chart_type_selector",
            format_func=lambda x: {
                "line": "Linjediagram",
                "heatmap": "Varmekart",
                "box": "Boksplot",
                "line_with_confidence": "Linje med konfidensintervall",
            }[x],
        )
        if not can_heatmap and not is_weekly:
            st.caption("Varmekart vises kun når du sammenligner minst 2 år (sesongmønster på tvers av år).")

    with col2:
        show_growth = st.checkbox("Vis vekstrater", value=False)

    if chart_type in {"heatmap", "box", "line_with_confidence"}:
        _render_altair_chart(df, point, chart_type)
    elif is_weekly:
        # Dedicated bar chart for weekly volume
        week_order = sorted(
            df["Week"].astype(str).tolist(),
            key=lambda s: int("".join([c for c in s if c.isdigit()]) or "0"),
        )
        fig = px.bar(
            df, x="Week", y="Volume",
            title=f"Ukentlig trafikkvolum for {point}",
            labels={"Volume": "Gjennomsnittlig døgntrafikk (ÅDT)", "Week": "Uke"},
            category_orders={"Week": week_order},
        )
        fig.update_traces(
            marker_color="#1f77b4",
            hovertemplate="%{x}: <b>%{y:,.0f}</b> ÅDT<extra></extra>",
        )
        fig.update_layout(
            template="plotly_white",
            height=480,
            yaxis=dict(tickformat=","),
            margin=dict(t=60, b=40),
        )
        st.plotly_chart(fig, use_container_width=True, key="weekly_bar", config={"displayModeBar": True})
    else:
        fig = create_advanced_visualization(df, point, chart_type)
        st.plotly_chart(
            fig,
            use_container_width=True,
            key=f"main_chart_{chart_type}",
            config={"displayModeBar": True, "responsive": True},
        )

    if show_growth:
        growth_df = calculate_growth_rates(df)
        growth_columns = [col for col in growth_df.columns if "Vekst" in col]
        if not growth_columns:
            st.info("Vekstrater krever minst 2 år i samme visning (bruk «Sammenlign år» og velg flere år).")
        else:
            st.subheader("Vekstrater (år-til-år)")
            if "Month Name" in growth_df.columns:
                id_vars = ["Month", "Month Name"]
                x_col = "Month Name"
            elif "Week" in growth_df.columns:
                id_vars = ["Week"]
                x_col = "Week"
            else:
                id_vars = []
                x_col = growth_df.index

            growth_melted = growth_df.melt(
                id_vars=id_vars,
                value_vars=growth_columns,
                var_name="Periode",
                value_name="Vekst (%)",
            )
            growth_melted["Vekst (%)"] = pd.to_numeric(growth_melted["Vekst (%)"], errors="coerce")
            growth_melted = growth_melted.dropna(subset=["Vekst (%)"])
            if growth_melted.empty:
                st.warning("Fant ingen gyldige vekstrater i datasettet (mangler baseline eller nullverdier).")
            else:
                zero = alt.Chart(pd.DataFrame({"y": [0]})).mark_rule(color="black", strokeDash=[4, 4]).encode(y="y:Q")
                bars = (
                    alt.Chart(growth_melted)
                    .mark_bar()
                    .encode(
                        x=alt.X(f"{x_col}:N", title="Periode"),
                        y=alt.Y("Vekst (%):Q", title="Vekst (%)"),
                        color=alt.condition(
                            alt.datum["Vekst (%)"] >= 0,
                            alt.value("#2ca02c"),
                            alt.value("#d62728"),
                        ),
                        tooltip=[alt.Tooltip("Periode:N"), alt.Tooltip("Vekst (%):Q", format="+.1f")],
                    )
                    .properties(height=320)
                )
                text = (
                    alt.Chart(growth_melted)
                    .mark_text(dy=alt.ExprRef("datum['Vekst (%)'] >= 0 ? -8 : 8"), fontSize=10)
                    .encode(
                        x=alt.X(f"{x_col}:N"),
                        y=alt.Y("Vekst (%):Q"),
                        text=alt.Text("Vekst (%):Q", format="+.1f"),
                        color=alt.condition(
                            alt.datum["Vekst (%)"] >= 0,
                            alt.value("#1a7a1a"),
                            alt.value("#a01a1a"),
                        ),
                    )
                )
                st.altair_chart((bars + zero + text).properties(title="År-til-år vekstrater"), use_container_width=True)


def process_data_for_years(
    point_ids: List[str], year_list: List[int], timeout_s: int, use_cache: bool, estimate_missing_points: bool
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    data: Dict[int, List[float]] = {}
    coverage_rows: List[pd.DataFrame] = []
    progress_bar = st.progress(0)
    status_text = st.empty()

    for i, year in enumerate(year_list):
        status_text.text(f"Henter data for {year}...")
        progress_bar.progress((i + 1) / len(year_list))
        traffic_data_dict = fetch_batch_traffic_data(point_ids, year, timeout_s, use_cache)
        if traffic_data_dict:
            monthly_sums, _, monthly_has_data, _ = sum_traffic_data(
                traffic_data_dict,
                expected_point_ids=point_ids,
                estimate_missing_points=estimate_missing_points,
            )
            data[year] = [v if has else np.nan for v, has in zip(monthly_sums, monthly_has_data)]
            coverage_rows.append(compute_monthly_coverage_summary(traffic_data_dict, year, point_ids))

    status_text.empty()
    progress_bar.empty()

    df = pd.DataFrame({"Month": list(range(1, 13))})
    for y in year_list:
        if y in data:
            df[str(y)] = data[y]
    df = add_month_names(df)
    numeric_columns = df.select_dtypes(include=[np.number]).columns
    df[numeric_columns] = df[numeric_columns].round(0).astype("Int64")
    coverage_df = pd.concat(coverage_rows, ignore_index=True) if coverage_rows else pd.DataFrame()
    return df, coverage_df


def process_data_for_months(
    point_ids: List[str], year: int, months: List[int], timeout_s: int, use_cache: bool, estimate_missing_points: bool
) -> Optional[Tuple[pd.DataFrame, pd.DataFrame]]:
    traffic_data_dict = fetch_batch_traffic_data(point_ids, year, timeout_s, use_cache)
    if not traffic_data_dict:
        return None
    data, _, monthly_has_data, _ = sum_traffic_data(
        traffic_data_dict,
        expected_point_ids=point_ids,
        estimate_missing_points=estimate_missing_points,
    )
    df = pd.DataFrame({"Month": list(range(1, 13)), str(year): data})
    df.loc[~pd.Series(monthly_has_data), str(year)] = np.nan
    df = df[df["Month"].isin(months)]
    df = add_month_names(df)
    numeric_columns = df.select_dtypes(include=[np.number]).columns
    df[numeric_columns] = df[numeric_columns].round(0).astype("Int64")
    coverage_df = compute_monthly_coverage_summary(traffic_data_dict, year, point_ids)
    coverage_df = coverage_df[coverage_df["month"].isin(months)].copy()
    return df, coverage_df


def process_data_for_weeks(
    point_ids: List[str], year: int, weeks: List[int], timeout_s: int, use_cache: bool
) -> Optional[Tuple[pd.DataFrame, pd.DataFrame]]:
    weekly_data_dict, weekly_cov_dict = fetch_weekly_traffic_data(point_ids, year, weeks, timeout_s, use_cache)
    if not weekly_data_dict:
        return None
    weekly_sums = sum_weekly_traffic_data(weekly_data_dict)
    df = pd.DataFrame([{"Week": week, "Volume": volume} for week, volume in weekly_sums.items()])
    df["Week_Num"] = df["Week"].str.extract(r"(\d+)").astype(int)
    df = df.sort_values("Week_Num").drop(columns=["Week_Num"]).reset_index(drop=True)
    df["Volume"] = pd.to_numeric(df["Volume"], errors="coerce").round(0).astype("Int64")
    coverage_df = compute_weekly_coverage_summary(weekly_data_dict, weekly_cov_dict, point_ids, year)
    return df, coverage_df


def render_data_coverage_banner(coverage_df: pd.DataFrame) -> None:
    if coverage_df is None or coverage_df.empty:
        return

    tmp = coverage_df.copy()
    tmp["has_issue"] = False
    if "points_expected" in tmp.columns and "points_present" in tmp.columns:
        tmp["has_issue"] |= tmp["points_present"] < tmp["points_expected"]
    if "mean_coverage_pct" in tmp.columns:
        tmp["has_issue"] |= tmp["mean_coverage_pct"].notna() & (
            tmp["mean_coverage_pct"] < (100.0 - FULL_COVERAGE_TOL_PCT)
        )
    if "min_coverage_pct" in tmp.columns:
        tmp["has_issue"] |= tmp["min_coverage_pct"].notna() & (tmp["min_coverage_pct"] < (100.0 - FULL_COVERAGE_TOL_PCT))

    issues = tmp[tmp["has_issue"]].copy()

    expected = int(tmp["points_expected"].dropna().iloc[0]) if "points_expected" in tmp.columns and tmp["points_expected"].notna().any() else 0
    present_min = int(tmp["points_present"].min()) if "points_present" in tmp.columns and tmp["points_present"].notna().any() else 0
    present_avg = float(tmp["points_present"].mean()) if "points_present" in tmp.columns and tmp["points_present"].notna().any() else float("nan")
    missing_periods = int((tmp["points_present"] < tmp["points_expected"]).sum()) if {"points_present", "points_expected"} <= set(tmp.columns) else 0
    periods_total = int(len(tmp))

    mean_cov = float(tmp["mean_coverage_pct"].mean()) if "mean_coverage_pct" in tmp.columns and tmp["mean_coverage_pct"].notna().any() else float("nan")
    min_cov = float(tmp["min_coverage_pct"].min()) if "min_coverage_pct" in tmp.columns and tmp["min_coverage_pct"].notna().any() else float("nan")

    point_cov_text = (
        f"**Punktdekning:** min {present_min}/{expected} punkter rapporterer"
        + (f" (snitt {present_avg:.1f}/{expected})" if expected and not pd.isna(present_avg) else "")
        + (f" — mangler i {missing_periods}/{periods_total} perioder" if periods_total else "")
    )
    data_cov_text = "**Datadekning:** "
    if not pd.isna(mean_cov):
        data_cov_text += f"snitt {mean_cov:.1f}%"
        if not pd.isna(min_cov):
            data_cov_text += f", min {min_cov:.1f}%"
        data_cov_text += " (på rapporterende punkter)"
    else:
        data_cov_text += "N/A"

    if issues.empty:
        st.success("✅ 100% i perioden som vises\n\n" + point_cov_text + "\n\n" + data_cov_text)
        return

    def _labels(sub: pd.DataFrame) -> List[str]:
        if "month_name" in sub.columns:
            months = [m for m in MONTH_NAMES if m in set(sub["month_name"].astype(str))]
            return months
        if "week" in sub.columns:
            return sub["week"].astype(str).tolist()
        return []

    if "year" in issues.columns and issues["year"].nunique() > 1:
        parts = [f"{int(y)}: " + ", ".join(_labels(sub)) for y, sub in issues.groupby("year")]
        where = " | ".join(parts)
    else:
        where = ", ".join(_labels(issues))

    st.warning("⚠️ Ikke 100% i perioden som vises\n\n" + point_cov_text + "\n\n" + data_cov_text + "\n\n" + f"**Perioder:** {where}")

    with st.expander("🔎 Detaljer (dekning)", expanded=False):
        try:
            if "month_name" in coverage_df.columns and "mean_coverage_pct" in coverage_df.columns:
                plot_df = coverage_df.dropna(subset=["mean_coverage_pct"]).copy()
                if not plot_df.empty:
                    plot_df["month_name"] = pd.Categorical(plot_df["month_name"], categories=MONTH_NAMES, ordered=True)
                    color = alt.Color("year:N", title="År") if "year" in plot_df.columns and plot_df["year"].nunique() > 1 else alt.value("#1f77b4")
                    tooltips = [
                        alt.Tooltip("month_name:N", title="Måned"),
                        alt.Tooltip("mean_coverage_pct:Q", format=".1f", title="Snitt dekning (%)"),
                    ]
                    if "year" in plot_df.columns:
                        tooltips.insert(0, alt.Tooltip("year:N", title="År"))
                    chart = (
                        alt.Chart(plot_df)
                        .mark_line(point=alt.OverlayMarkDef(size=50))
                        .encode(
                            x=alt.X("month_name:N", sort=MONTH_NAMES, title="Måned"),
                            y=alt.Y("mean_coverage_pct:Q", title="Snitt dekning (%)", scale=alt.Scale(domain=[0, 100])),
                            color=color,
                            tooltip=tooltips,
                        )
                        .properties(height=220)
                    )
                    st.altair_chart(chart, use_container_width=True)
            elif "week" in coverage_df.columns and "mean_coverage_pct" in coverage_df.columns:
                plot_df = coverage_df.dropna(subset=["mean_coverage_pct"]).copy()
                if not plot_df.empty:
                    week_order = sorted(plot_df["week"].astype(str).unique(), key=lambda s: int("".join([c for c in s if c.isdigit()]) or "0"))
                    chart = (
                        alt.Chart(plot_df)
                        .mark_line(point=alt.OverlayMarkDef(size=50), color="#1f77b4")
                        .encode(
                            x=alt.X("week:N", sort=week_order, title="Uke"),
                            y=alt.Y("mean_coverage_pct:Q", title="Snitt dekning (%)", scale=alt.Scale(domain=[0, 100])),
                            tooltip=[
                                alt.Tooltip("week:N", title="Uke"),
                                alt.Tooltip("mean_coverage_pct:Q", format=".1f", title="Snitt dekning (%)"),
                            ],
                        )
                        .properties(height=220)
                    )
                    st.altair_chart(chart, use_container_width=True)
        except Exception as exc:
            logger.warning("Dekning: klarte ikke vise detalj-graf: %s", exc)

        cols: List[str] = []
        if "year" in coverage_df.columns:
            cols.append("year")
        if "month_name" in coverage_df.columns:
            cols.append("month_name")
        if "week" in coverage_df.columns:
            cols.append("week")
        cols += [c for c in ["points_present", "points_expected", "mean_coverage_pct", "min_coverage_pct"] if c in coverage_df.columns]
        view = coverage_df[cols].copy()
        for c in ["mean_coverage_pct", "min_coverage_pct", "points_present_pct"]:
            if c in view.columns:
                view[c] = pd.to_numeric(view[c], errors="coerce").round(1)
        st.dataframe(view, use_container_width=True, hide_index=True)


def render_anomaly_banner(df: pd.DataFrame) -> None:
    anomalies = detect_monthly_anomalies(df, threshold_pct=float(ANOMALY_THRESHOLD_PCT))
    if anomalies is None or anomalies.empty:
        return
    preview = anomalies.sort_values(["year", "month"]).head(6)
    items = [f"{int(r.year)} {str(r.month_name)} ({float(r.deviation_pct):+.0f}%)" for r in preview.itertuples(index=False)]
    st.warning("⚠️ Mulige anomalier: " + ", ".join(items) + (" …" if len(anomalies) > len(preview) else ""))


def render_point_basis_note(coverage_df: pd.DataFrame) -> None:
    if coverage_df is None or coverage_df.empty:
        return
    if not {"points_present", "points_expected"} <= set(coverage_df.columns):
        return
    expected = int(coverage_df["points_expected"].dropna().iloc[0]) if coverage_df["points_expected"].notna().any() else 0
    if not expected:
        return
    present_min = int(coverage_df["points_present"].min()) if coverage_df["points_present"].notna().any() else 0
    if present_min < expected:
        st.caption(f"⚠️ Grafene kan være basert på ufullstendige målepunkter (min {present_min}/{expected} rapporterer i perioden).")


def render_totals_tab(df: pd.DataFrame, point: str, comparison_mode: str, year_list: List[int], year: int, point_ids: List[str], timeout_s: int, use_cache: bool):
    st.subheader("🧮 Totale passeringer (samlet trafikk)")
    if comparison_mode == COMPARE_WEEKS:
        st.info(
            "Ukesvisning viser per nå gjennomsnitt per døgn (sum av punkter). "
            "Totale uketall krever at vi vet hvor mange gyldige dager som inngår i snittet."
        )
        return

    years_for_totals = year_list if comparison_mode == COMPARE_YEARS else [year]
    totals_df = compute_monthly_totals_table(df, years_for_totals)
    st.caption("Beregning: (månedsvis gjennomsnitt per døgn) × (antall dager i måneden), summert per måned.")

    year_cols = [y for y in years_for_totals if str(y) in totals_df.columns]
    if year_cols:
        latest_year = max(year_cols)
        total, months_present, days_covered = calculate_yearly_total_from_monthly_averages(df, int(latest_year))
        avg_per_day = (total / days_covered) if days_covered else None
        full_year_estimate = (avg_per_day * days_in_year(int(latest_year))) if avg_per_day is not None else None

        m1, m2, m3 = st.columns(3)
        with m1:
            st.metric(
                f"Totalt {latest_year} ({'hittil' if months_present < 12 else 'helår'})",
                format_number(total),
                help="Totalt antall passeringer beregnet fra Vegvesen sine måneds-ÅDT.",
            )
        with m2:
            st.metric("Snitt per døgn (basert på tilgjengelige måneder)", format_number(avg_per_day) if avg_per_day is not None else "N/A")
        with m3:
            if months_present < 12 and full_year_estimate is not None:
                st.metric("Estimert helår", format_number(full_year_estimate))
            else:
                st.metric("Måneder med data", f"{months_present}/12")

        if comparison_mode == COMPARE_YEARS and len(year_cols) >= 2:
            prev_year = sorted(year_cols)[-2]
            prev_total, _, _ = calculate_yearly_total_from_monthly_averages(df, int(prev_year))
            if prev_total:
                st.metric(f"Endring vs {prev_year}", f"{((total - prev_total) / prev_total * 100):+.1f}%")

    melt_cols = [str(y) for y in years_for_totals if str(y) in totals_df.columns]
    if melt_cols:
        melted = totals_df.melt(id_vars=["Month", "Month Name"], value_vars=melt_cols, var_name="År", value_name="Passeringer")
        fig_totals = px.bar(
            melted, x="Month Name", y="Passeringer", color="År", barmode="group",
            title="Totale passeringer per måned",
            labels={"Passeringer": "Passeringer", "Month Name": "Måned"},
        )
        fig_totals.update_traces(hovertemplate="%{x}: <b>%{y:,.0f}</b><extra>%{fullData.name}</extra>")
        fig_totals.update_layout(
            yaxis=dict(tickformat=","),
            legend=dict(orientation="h", yanchor="bottom", y=1.02),
            margin=dict(t=60, b=40),
        )
        st.plotly_chart(fig_totals, use_container_width=True)

    if point == "Ryfast (sum tunneler)":
        with st.expander("🔎 Fordeling mellom tunneler (Ryfylketunnelen vs Hundvågtunnelen)"):
            breakdown_year = st.selectbox("Velg år for fordeling", options=years_for_totals, index=len(years_for_totals) - 1)
            traffic_by_point = fetch_batch_traffic_data(point_ids, int(breakdown_year), timeout_s, use_cache)
            point_ids_by_group = {
                "Ryfylketunnelen": TRAFFIC_POINTS["Ryfylketunnelen"]["ids"],
                "Hundvågtunnelen": (
                    TRAFFIC_POINTS["Hundvågtunnelen"]["ids"]
                    if st.session_state.get("ryfast_include_ramp", True)
                    else HUNDVAG_TUNNEL_IDS_UTEN_PÅRAMPE
                ),
            }
            totals_by_group_df, coverage_by_group_df = aggregate_monthly_totals_by_group(traffic_by_point, point_ids_by_group, int(breakdown_year))
            melted_groups = totals_by_group_df.melt(id_vars=["Month", "Month Name"], value_vars=list(point_ids_by_group.keys()), var_name="Tunnel", value_name="Passeringer")
            fig_stack = px.bar(melted_groups, x="Month Name", y="Passeringer", color="Tunnel", title=f"Samlet trafikk {breakdown_year} (stacked per tunnel)")
            fig_stack.update_yaxes(tickformat=",")
            st.plotly_chart(fig_stack, use_container_width=True)
            st.caption("Dekning (%) er snitt av rapportert dekning på målepunktene per måned.")
            st.dataframe(coverage_by_group_df.round(1), use_container_width=True, hide_index=True)

    with st.expander("📋 Totaltabell"):
        formatted_totals = totals_df.copy()
        numeric_cols = formatted_totals.select_dtypes(include=[np.number]).columns
        for col in numeric_cols:
            formatted_totals[col] = formatted_totals[col].map(format_number)
        st.dataframe(formatted_totals, use_container_width=True, hide_index=True)


def render_data_quality_tab(
    point: str,
    comparison_mode: str,
    year_list: List[int],
    year: int,
    point_ids: List[str],
    timeout_s: int,
    use_cache: bool,
    coverage_threshold: float,
    df: Optional[pd.DataFrame] = None,
    coverage_summary: Optional[pd.DataFrame] = None,
):
    st.subheader("🛡️ Dekning og datakvalitet")
    st.caption(
        "Dekning (%) er rapportert datadekning fra Vegvesen. "
        "Lav dekning kan gi skjevheter. Usikkerhet er indikativ og basert på oppgitte konfidensintervaller."
    )

    if comparison_mode == COMPARE_WEEKS:
        cov = coverage_summary if coverage_summary is not None else pd.DataFrame()
        if cov is None or cov.empty:
            st.warning("Fant ingen dekningsdata for ukesvisning.")
            return

        below_cov = cov[(cov["mean_coverage_pct"].notna()) & (cov["mean_coverage_pct"] < float(coverage_threshold))]
        m1, m2, m3 = st.columns(3)
        with m1:
            st.metric("Dekningsterskel", f"{coverage_threshold:.0f}%")
        with m2:
            st.metric("Uker under terskel", str(len(below_cov)))
        with m3:
            st.metric(
                "Snitt dekning",
                f"{cov['mean_coverage_pct'].mean():.1f}%" if cov["mean_coverage_pct"].notna().any() else "N/A",
            )

        st.markdown("### Ukentlig dekning")
        try:
            plot_df = cov.dropna(subset=["mean_coverage_pct"]).copy()
            if not plot_df.empty:
                week_order = sorted(
                    plot_df["week"].astype(str).unique(),
                    key=lambda s: int("".join([c for c in s if c.isdigit()]) or "0"),
                )
                chart = (
                    alt.Chart(plot_df)
                    .mark_line(point=True, color="#1f77b4")
                    .encode(
                        x=alt.X("week:N", sort=week_order, title="Uke"),
                        y=alt.Y("mean_coverage_pct:Q", title="Snitt dekning (%)", scale=alt.Scale(domain=[0, 100])),
                        tooltip=[
                            alt.Tooltip("week:N", title="Uke"),
                            alt.Tooltip("points_present:Q", title="Punkter m/data"),
                            alt.Tooltip("points_expected:Q", title="Punkter forventet"),
                            alt.Tooltip("mean_coverage_pct:Q", format=".1f", title="Snitt dekning (%)"),
                            alt.Tooltip("min_coverage_pct:Q", format=".1f", title="Min dekning (%)"),
                        ],
                    )
                    .properties(height=320)
                )
                st.altair_chart(chart, use_container_width=True)
        except Exception as exc:
            logger.warning("Datakvalitet (uker): klarte ikke vise dekningsgraf: %s", exc)

        st.dataframe(cov.copy(), use_container_width=True, hide_index=True)
        return

    years = year_list if comparison_mode == COMPARE_YEARS else [year]
    selected_year = st.selectbox("År", options=years, index=len(years) - 1, key="dq_year")

    traffic_by_point = fetch_batch_traffic_data(point_ids, int(selected_year), timeout_s, use_cache)
    metrics = extract_point_monthly_metrics(traffic_by_point, int(selected_year))
    if metrics.empty:
        st.warning("Fant ingen dekning/data for valgt år.")
        return

    below = metrics[(metrics["coverage_pct"].notna()) & (metrics["coverage_pct"] < coverage_threshold)]
    m1, m2, m3 = st.columns(3)
    with m1:
        st.metric("Dekningsterskel", f"{coverage_threshold:.0f}%")
    with m2:
        st.metric("Observasjoner under terskel", str(len(below)))
    with m3:
        st.metric("Snitt dekning", f"{metrics['coverage_pct'].mean():.1f}%" if metrics["coverage_pct"].notna().any() else "N/A")

    if not below.empty:
        bad_months = (
            below.groupby("month_name")["point_label"]
            .apply(lambda s: ", ".join(sorted(set(s))[:4]) + (" …" if len(set(s)) > 4 else ""))
            .reindex(MONTH_NAMES)
            .dropna()
        )
        st.warning("Måneder med lav dekning: " + ", ".join([m for m in bad_months.index.tolist() if m]))

    # Coverage per point heatmap/table
    st.markdown("### Dekning per målepunkt")
    cov_piv = coverage_pivot(metrics)
    if not cov_piv.empty:
        cov_long = cov_piv.reset_index().melt(id_vars=["month_name"], var_name="Målepunkt", value_name="Dekning")
        cov_long = cov_long.dropna(subset=["Dekning"])
        chart = (
            alt.Chart(cov_long)
            .mark_rect()
            .encode(
                x=alt.X("month_name:N", sort=MONTH_NAMES, title="Måned"),
                y=alt.Y("Målepunkt:N", title="Målepunkt"),
                color=alt.Color("Dekning:Q", scale=alt.Scale(domain=[0, 100], scheme="yellowgreenblue"), title="Dekning (%)"),
                tooltip=[
                    alt.Tooltip("Målepunkt:N"),
                    alt.Tooltip("month_name:N", title="Måned"),
                    alt.Tooltip("Dekning:Q", format=".1f", title="Dekning (%)"),
                ],
            )
            .properties(height=min(420, 24 * max(1, len(cov_piv.columns))), title="Dekning per måned og målepunkt")
        )
        st.altair_chart(chart, use_container_width=True)

        with st.expander("📋 Dekningstabell"):
            st.dataframe(cov_piv.round(1), use_container_width=True)

    # Group coverage (Ryfast breakdown)
    if point == "Ryfast (sum tunneler)":
        st.markdown("### Dekning per tunnel")
        point_ids_by_group = {
            "Ryfylketunnelen": TRAFFIC_POINTS["Ryfylketunnelen"]["ids"],
            "Hundvågtunnelen": (
                TRAFFIC_POINTS["Hundvågtunnelen"]["ids"]
                if st.session_state.get("ryfast_include_ramp", True)
                else HUNDVAG_TUNNEL_IDS_UTEN_PÅRAMPE
            ),
        }
        grouped = group_coverage_by_month(metrics, point_ids_by_group)
        if not grouped.empty:
            chart = (
                alt.Chart(grouped)
                .mark_line(point=True)
                .encode(
                    x=alt.X("month_name:N", sort=MONTH_NAMES, title="Måned"),
                    y=alt.Y("coverage_pct:Q", title="Dekning (%)", scale=alt.Scale(domain=[0, 100])),
                    color=alt.Color("group:N", title="Tunnel"),
                    tooltip=[
                        alt.Tooltip("group:N", title="Tunnel"),
                        alt.Tooltip("month_name:N", title="Måned"),
                        alt.Tooltip("coverage_pct:Q", format=".1f", title="Dekning (%)"),
                    ],
                )
                .properties(height=320)
            )
            st.altair_chart(chart, use_container_width=True)

    st.markdown("### Usikkerhet (indikativ)")
    st.caption(
        "Konfidensintervaller er oppgitt per målepunkt. Vi summerer bounds på tvers av punkt for å gi et "
        "indikativt spenn (ikke et strengt statistisk intervall)."
    )
    totals_ci = totals_with_uncertainty_from_metrics(metrics)
    if not totals_ci.empty:
        totals_ci["total"] = totals_ci["total"].round(0).astype("Int64")
        totals_ci["total_lower"] = totals_ci["total_lower"].round(0).astype("Int64")
        totals_ci["total_upper"] = totals_ci["total_upper"].round(0).astype("Int64")
        display = totals_ci[["month_name", "total", "total_lower", "total_upper", "coverage_pct"]].copy()
        display["coverage_pct"] = display["coverage_pct"].round(1)
        display.columns = ["Måned", "Totalt", "Nedre", "Øvre", "Dekning (%)"]
        st.dataframe(display, use_container_width=True, hide_index=True)

    if df is not None and not df.empty:
        st.markdown("### Anomali-varsler (indikativ)")
        st.caption(
            f"Flagger måneder der nivået avviker mer enn ±{ANOMALY_THRESHOLD_PCT:.0f}% fra forventet basert på de andre valgte årene."
        )
        anomalies = detect_monthly_anomalies(df, threshold_pct=float(ANOMALY_THRESHOLD_PCT))
        if anomalies.empty:
            st.caption("Ingen anomalier flagget.")
        else:
            st.warning(f"Flagget {len(anomalies)} anomalier.")
            st.dataframe(
                anomalies.assign(deviation_pct=anomalies["deviation_pct"].round(1)),
                use_container_width=True,
                hide_index=True,
            )


def main():
    init_session_state()

    st.set_page_config(page_title="Trafikkdata Visualisering - Ryfast", page_icon="🚗", layout="wide", initial_sidebar_state="expanded")
    st.markdown(
        """
    <style>
    .main-header { background: linear-gradient(90deg, #1f77b4, #ff7f0e); color: white; padding: 1rem;
      border-radius: 10px; text-align: center; margin-bottom: 2rem; }
    </style>
    """,
        unsafe_allow_html=True,
    )
    st.markdown(
        """
    <div class="main-header">
        <h1>🚗 Trafikkdata Visualisering - Ryfast</h1>
        <p>Analyse av trafikkmønstre for Ryfylketunnelen, Hundvågtunnelen og Bybrua</p>
    </div>
    """,
        unsafe_allow_html=True,
    )

    st.sidebar.header("⚙️ Innstillinger")
    point_options = list(TRAFFIC_POINTS.keys())
    point_descriptions = [f"{p} - {TRAFFIC_POINTS[p]['description']}" for p in point_options]

    with st.sidebar.form("controls", clear_on_submit=False):
        selected_index = st.selectbox(
            "Velg målepunkt",
            range(len(point_options)),
            format_func=lambda x: point_descriptions[x],
            key="point_selector",
        )
        point = point_options[selected_index]

        with st.expander("ℹ️ Om valgt målepunkt"):
            st.write(f"**Beskrivelse:** {TRAFFIC_POINTS[point]['description']}")
            st.write(f"**Åpnet:** {TRAFFIC_POINTS[point]['opened']}")

        comparison_mode = st.radio(
            "Velg analysetype",
            [COMPARE_YEARS, COMPARE_MONTHS, COMPARE_WEEKS],
            key="comparison_mode",
        )

        with st.expander("🔧 Avanserte innstillinger"):
            use_cache = st.checkbox("Bruk hurtigbuffer", value=True, key="use_cache")
            timeout_s = st.slider("API timeout (sekunder)", 10, 90, 60, key="timeout_s")
            coverage_threshold = st.slider("Min. dekning (%)", 50, 100, 90, key="coverage_threshold")
            estimate_missing_points = st.checkbox(
                "Estimer manglende målepunkter (pro-rata)",
                value=False,
                key="estimate_missing_points",
                help="Når noen (men ikke alle) målepunkter mangler data i en måned, skaleres totalen opp basert på antall punkter med data. Bruk med varsomhet.",
            )

        direction = "Begge retninger"
        has_optional_points = point in {"Ryfast (sum tunneler)", "Hundvågtunnelen"}
        include_ramp = st.checkbox(
            "Inkluder pårampe/tilleggspunkt (der definert)",
            value=bool(st.session_state.get("ryfast_include_ramp", True)),
            key="ryfast_include_ramp",
            disabled=not has_optional_points,
            help=(
                "Når av (der tilgjengelig), ekskluderes pårampe/tilleggspunkt for Hundvågtunnelen "
                f"({', '.join(HUNDVAG_TUNNEL_RAMP_IDS)})."
            ),
        )
        if point == "Bybrua":
            direction = st.selectbox("Velg retning", ["Begge retninger", "Mot nord", "Mot sør"], key="bybrua_direction")

        year_input = DEFAULT_YEARS
        year_list: List[int] = []
        year = 2025
        months: List[int] = []
        weeks: List[int] = []

        if comparison_mode == COMPARE_YEARS:
            year_input = st.text_input("År (komma-separert)", value=DEFAULT_YEARS, key="years_input")
        elif comparison_mode == COMPARE_MONTHS:
            year = st.selectbox("År", list(YEAR_RANGE), index=list(YEAR_RANGE).index(2025) if 2025 in YEAR_RANGE else 0, key="months_year")
            months = st.multiselect(
                "Velg måneder",
                list(range(1, 13)),
                default=list(range(1, 13)),
                format_func=lambda m: MONTH_NAMES[m - 1],
                key="months_selected",
            )
        else:
            year = st.selectbox("År", list(YEAR_RANGE), index=list(YEAR_RANGE).index(2025) if 2025 in YEAR_RANGE else 0, key="weeks_year")
            weeks = st.multiselect("Velg uker", list(range(1, 53)), default=list(range(1, 11)), key="weeks_selected")

        submitted = st.form_submit_button("📊 Analyser data", type="primary")

    render_api_status_sidebar()

    if st.sidebar.button("🗑️ Tøm cache"):
        st.cache_data.clear()
        st.session_state.last_result = None
        st.sidebar.success("Cache tømt!")

    def resolve_point_ids() -> List[str]:
        if point == "Ryfast (sum tunneler)":
            ryfylke_ids = TRAFFIC_POINTS["Ryfylketunnelen"]["ids"]
            hundvag_ids = TRAFFIC_POINTS["Hundvågtunnelen"]["ids"] if include_ramp else HUNDVAG_TUNNEL_IDS_UTEN_PÅRAMPE
            return ryfylke_ids + hundvag_ids
        if point == "Ryfylketunnelen":
            return TRAFFIC_POINTS["Ryfylketunnelen"]["ids"]
        if point == "Hundvågtunnelen":
            return TRAFFIC_POINTS["Hundvågtunnelen"]["ids"] if include_ramp else HUNDVAG_TUNNEL_IDS_UTEN_PÅRAMPE
        if direction == "Begge retninger":
            return TRAFFIC_POINTS["Bybrua"]["ids"]["Mot nord"] + TRAFFIC_POINTS["Bybrua"]["ids"]["Mot sør"]
        return TRAFFIC_POINTS["Bybrua"]["ids"][direction]

    if submitted:
        if comparison_mode == COMPARE_YEARS:
            try:
                year_list = [int(y.strip()) for y in year_input.split(",") if y.strip()]
            except Exception:
                st.sidebar.error("Ugyldig format. Bruk f.eks. 2024,2025")
                st.session_state.last_result = None
                year_list = []
            if not year_list:
                st.sidebar.warning("Velg minst ett år")
                st.session_state.last_result = None
            else:
                year = year_list[-1]

        if comparison_mode == COMPARE_MONTHS and not months:
            st.sidebar.warning("Velg minst én måned")
            st.session_state.last_result = None
        if comparison_mode == COMPARE_WEEKS and not weeks:
            st.sidebar.warning("Velg minst én uke")
            st.session_state.last_result = None

        point_ids = resolve_point_ids()

        if (comparison_mode == COMPARE_YEARS and year_list) or (comparison_mode != COMPARE_YEARS):
            with st.spinner("🔄 Behandler data..."):
                if comparison_mode == COMPARE_YEARS:
                    result_tuple = process_data_for_years(point_ids, year_list, timeout_s, use_cache, estimate_missing_points)
                    df, coverage_summary = result_tuple
                    title = f"Årlig sammenligning for {point}"
                elif comparison_mode == COMPARE_MONTHS:
                    result_tuple = process_data_for_months(point_ids, year, months, timeout_s, use_cache, estimate_missing_points)
                    if result_tuple is None:
                        df, coverage_summary = None, pd.DataFrame()
                    else:
                        df, coverage_summary = result_tuple
                    title = f"Månedlig analyse for {point} i {year}"
                else:
                    result_tuple = process_data_for_weeks(point_ids, year, weeks, timeout_s, use_cache)
                    if result_tuple is None:
                        df, coverage_summary = None, pd.DataFrame()
                    else:
                        df, coverage_summary = result_tuple
                    title = f"Ukentlig analyse for {point} i {year}"

            if df is None or df.empty:
                st.error("❌ Ingen data tilgjengelig for valgte kriterier")
                st.session_state.last_result = None
            else:
                st.session_state.last_result = {
                    "df": df,
                    "coverage_summary": coverage_summary,
                    "title": title,
                    "point": point,
                    "comparison_mode": comparison_mode,
                    "year_list": year_list,
                    "year": year,
                    "point_ids": point_ids,
                    "timeout_s": timeout_s,
                    "use_cache": use_cache,
                    "coverage_threshold": coverage_threshold,
                    "estimate_missing_points": bool(estimate_missing_points),
                }

    if not st.session_state.last_result:
        st.info("Velg målepunkt og analysetype i sidebaren, og trykk «Analyser data».")
        return

    result = st.session_state.last_result
    df = result["df"]
    coverage_summary = result.get("coverage_summary", pd.DataFrame())
    title = result["title"]
    point = result["point"]
    comparison_mode = result["comparison_mode"]
    year_list = result["year_list"]
    year = result["year"]
    point_ids = result["point_ids"]
    timeout_s = result["timeout_s"]
    use_cache = result["use_cache"]
    coverage_threshold = result.get("coverage_threshold", 90)

    tab1, tab2, tab3, tab4, tab5 = st.tabs(["📈 Visualisering", "📊 Data", "🧮 Totaltall", "🛡️ Datakvalitet", "📄 Rapport"])

    with tab1:
        st.subheader(title)
        render_data_coverage_banner(coverage_summary)
        render_anomaly_banner(df)
        create_comparison_dashboard(df, point)
        render_point_basis_note(coverage_summary)

    with tab2:
        st.subheader("📊 Rådata")
        formatted_df = df.copy()
        numeric_cols = formatted_df.select_dtypes(include=[np.number]).columns
        for col in numeric_cols:
            formatted_df[col] = formatted_df[col].map(format_number)
        st.dataframe(formatted_df, use_container_width=True, hide_index=True)
        if coverage_summary is not None and not coverage_summary.empty:
            with st.expander("🛡️ Datadekning (for perioden som vises)"):
                cov_view = coverage_summary.copy()
                for c in ["points_present_pct", "mean_coverage_pct", "min_coverage_pct"]:
                    if c in cov_view.columns:
                        cov_view[c] = pd.to_numeric(cov_view[c], errors="coerce").round(1)
                st.dataframe(cov_view, use_container_width=True, hide_index=True)

    with tab3:
        render_totals_tab(df, point, comparison_mode, year_list, year, point_ids, timeout_s, use_cache)

    with tab4:
        render_data_quality_tab(
            point=point,
            comparison_mode=comparison_mode,
            year_list=year_list,
            year=year,
            point_ids=point_ids,
            timeout_s=timeout_s,
            use_cache=use_cache,
            coverage_threshold=float(coverage_threshold),
            df=df,
            coverage_summary=coverage_summary,
        )

    with tab5:
        st.subheader("📄 Rapport / eksport")
        create_export_section(df, point, coverage_summary=coverage_summary)


if __name__ == "__main__":
    main()
