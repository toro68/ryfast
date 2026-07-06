"""Excel-eksport av trafikkdata-rapporter (openpyxl)."""

import io
import logging
from datetime import datetime
from typing import List, Optional, Tuple

import pandas as pd
import streamlit as st

from ryfast_app.api import fetch_batch_traffic_data
from ryfast_app.config import (
    ANOMALY_THRESHOLD_PCT,
    COMPARE_WEEKS,
    COMPARE_YEARS,
    HUNDVAG_TUNNEL_IDS_UTEN_PÅRAMPE,
    MONTH_NAMES,
    TRAFFIC_POINTS,
)
from ryfast_app.metrics import (
    aggregate_monthly_totals_by_group,
    calculate_growth_rates,
    calculate_seasonal_patterns,
    compute_monthly_coverage_summary,
    compute_monthly_totals_table,
    coverage_pivot,
    group_coverage_by_month,
)
from ryfast_app.processing import (
    calculate_yearly_total_from_monthly_averages,
    days_in_year,
    detect_monthly_anomalies,
    extract_point_monthly_metrics,
)

logger = logging.getLogger(__name__)


def _excel_row(row) -> list:
    """openpyxl godtar ikke pd.NA/NaN; erstatt med None (tom celle)."""
    return [None if pd.isna(value) else value for value in row]


try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Alignment, Font, PatternFill

    OPENPYXL_AVAILABLE = True
except ImportError:
    Workbook = None  # type: ignore[assignment]
    BarChart = None  # type: ignore[assignment]
    LineChart = None  # type: ignore[assignment]
    Reference = None  # type: ignore[assignment]
    CellIsRule = None  # type: ignore[assignment]
    Alignment = None  # type: ignore[assignment]
    Font = None  # type: ignore[assignment]
    PatternFill = None  # type: ignore[assignment]
    OPENPYXL_AVAILABLE = False


def export_to_excel(df: pd.DataFrame) -> bytes:
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is not available")
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Trafikkdata", index=False)
        if len([col for col in df.columns if str(col).isdigit()]) >= 2:
            calculate_growth_rates(df).to_excel(writer, sheet_name="Vekstrater", index=False)
        seasonal = calculate_seasonal_patterns(df)
        if seasonal:
            pd.DataFrame(seasonal).T.to_excel(writer, sheet_name="Sesongmønstre")
    return output.getvalue()

def build_excel_report(
    df: pd.DataFrame,
    point: str,
    comparison_mode: str,
    year_list: List[int],
    year: int,
    point_ids: List[str],
    timeout_s: int,
    use_cache: bool,
    coverage_threshold: float,
    ryfast_include_ramp: Optional[bool] = None,
    estimate_missing_points: bool = False,
) -> bytes:
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is not available")

    include_ramp = (
        bool(st.session_state.get("ryfast_include_ramp", True)) if ryfast_include_ramp is None else bool(ryfast_include_ramp)
    )

    header_font = Font(bold=True)
    title_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
    wrap_top = Alignment(vertical="top", wrap_text=True)
    num_format_int = "#,##0"
    num_format_pct_1 = "0.0"

    def style_header_row(ws, row: int = 1) -> None:
        for cell in ws[row]:
            cell.font = header_font
            cell.fill = header_fill

    wb = Workbook()

    # Summary (first sheet)
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary["A1"] = "Ryfast - rapport"
    ws_summary["A1"].font = title_font
    ws_summary["A2"] = f"Målepunkt: {point}"
    ws_summary["A3"] = f"Analysetype: {comparison_mode}"
    ws_summary["A4"] = f"År: {','.join(map(str, year_list)) if comparison_mode == COMPARE_YEARS else str(year)}"
    ws_summary["A5"] = f"Generert: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_summary["A6"] = f"Min. dekning: {coverage_threshold:.0f}%"
    for r in range(2, 7):
        ws_summary[f"A{r}"].alignment = wrap_top
    ws_summary.column_dimensions["A"].width = 80

    # Metadata
    ws_meta = wb.create_sheet("Metadata")
    ws_meta["A1"] = "Metadata"
    ws_meta["A1"].font = title_font
    meta_rows = [
        ("Generert", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Målepunkt", point),
        ("Analysetype", comparison_mode),
        ("År", ",".join(map(str, year_list)) if comparison_mode == COMPARE_YEARS else str(year)),
        ("Punkt-IDer", ", ".join(point_ids)),
        ("Timeout (s)", str(timeout_s)),
        ("Cache", "Ja" if use_cache else "Nei"),
        ("Min. dekning (%)", f"{coverage_threshold:.0f}"),
        ("Estimer manglende punkt", "Ja" if estimate_missing_points else "Nei"),
    ]
    for i, (k, v) in enumerate(meta_rows, start=3):
        ws_meta[f"A{i}"] = k
        ws_meta[f"B{i}"] = v
        ws_meta[f"A{i}"].font = header_font
        ws_meta[f"B{i}"].alignment = wrap_top
    ws_meta.column_dimensions["A"].width = 20
    ws_meta.column_dimensions["B"].width = 120

    # Summary - yearly totals from df (no extra API calls)
    year_cols_in_df = sorted([int(c) for c in df.columns if str(c).isdigit()])
    if comparison_mode != COMPARE_WEEKS and year_cols_in_df:
        ws_summary["A8"] = "Årsoppsummering (beregnet fra månedsdata)"
        ws_summary["A8"].font = Font(bold=True, size=12)
        ws_summary.append(["År", "Sum (mnd)", "Måneder", "Dager dekket", "Snitt per døgn", "Helårsestimat", "YoY (%)"])
        style_header_row(ws_summary, row=9)

        prev_estimate: Optional[float] = None
        for y in year_cols_in_df:
            total, months_present, days_covered = calculate_yearly_total_from_monthly_averages(df, int(y))
            avg_per_day = (total / days_covered) if days_covered else None
            helar = None
            if avg_per_day is not None and not pd.isna(avg_per_day):
                helar = float(avg_per_day) * days_in_year(int(y))
            yoy = None
            if prev_estimate and helar and prev_estimate != 0:
                yoy = (helar - prev_estimate) / prev_estimate * 100
            ws_summary.append(
                [
                    int(y),
                    float(total) if total is not None else None,
                    int(months_present),
                    int(days_covered),
                    float(avg_per_day) if avg_per_day is not None else None,
                    float(helar) if helar is not None else None,
                    float(yoy) if yoy is not None else None,
                ]
            )
            prev_estimate = helar if helar is not None else prev_estimate

        first_data_row = 10
        last_data_row = 9 + len(year_cols_in_df)
        for r in range(first_data_row, last_data_row + 1):
            ws_summary.cell(row=r, column=2).number_format = num_format_int
            ws_summary.cell(row=r, column=5).number_format = num_format_int
            ws_summary.cell(row=r, column=6).number_format = num_format_int
            ws_summary.cell(row=r, column=7).number_format = num_format_pct_1

        for col, width in [("B", 14), ("C", 10), ("D", 12), ("E", 14), ("F", 14), ("G", 10)]:
            ws_summary.column_dimensions[col].width = width

        try:
            chart = BarChart()
            chart.title = "Helårsestimat (beregnet)"
            chart.y_axis.title = "Passeringer"
            data_ref = Reference(ws_summary, min_col=6, min_row=9, max_row=last_data_row)
            cats = Reference(ws_summary, min_col=1, min_row=10, max_row=last_data_row)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 9
            chart.width = 20
            ws_summary.add_chart(chart, "I8")
        except Exception as exc:
            logger.warning("Excel: klarte ikke legge til årsestimatgraf: %s", exc)

    # Data sheet
    ws_data = wb.create_sheet("Data")
    ws_data.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws_data.append(_excel_row(row))
    ws_data.freeze_panes = "A2"
    ws_data.auto_filter.ref = ws_data.dimensions
    style_header_row(ws_data, row=1)

    # Totals sheet (if applicable)
    if comparison_mode != COMPARE_WEEKS:
        years_for_totals = year_list if comparison_mode == COMPARE_YEARS else [year]
        totals_df = compute_monthly_totals_table(df, years_for_totals)
        ws_tot = wb.create_sheet("Totals")
        ws_tot.append(list(totals_df.columns))
        for row in totals_df.itertuples(index=False):
            ws_tot.append(_excel_row(row))
        ws_tot.freeze_panes = "A2"
        ws_tot.auto_filter.ref = ws_tot.dimensions
        style_header_row(ws_tot, row=1)
        for col_idx, col_name in enumerate(totals_df.columns, start=1):
            if str(col_name).isdigit():
                for r in range(2, 2 + len(totals_df)):
                    ws_tot.cell(row=r, column=col_idx).number_format = num_format_int

        # Simple chart
        if len(years_for_totals) >= 1:
            chart = LineChart()
            chart.title = "Totale passeringer per måned"
            chart.y_axis.title = "Passeringer"
            chart.x_axis.title = "Måned"
            data_ref = Reference(
                ws_tot,
                min_col=3,
                min_row=1,
                max_col=2 + len(years_for_totals),
                max_row=1 + len(totals_df),
            )
            cats = Reference(ws_tot, min_col=2, min_row=2, max_row=1 + len(totals_df))
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)
            ws_tot.add_chart(chart, "H2")

    # Coverage sheets (on-demand for a selected year)
    if comparison_mode != COMPARE_WEEKS:
        selected_year = max(year_list) if comparison_mode == COMPARE_YEARS else year
        traffic_by_point = fetch_batch_traffic_data(point_ids, int(selected_year), timeout_s, use_cache)
        metrics = extract_point_monthly_metrics(traffic_by_point, int(selected_year))
        if not metrics.empty:
            cov_piv = coverage_pivot(metrics)
            ws_cov = wb.create_sheet("Coverage (per point)")
            ws_cov.append(["Måned"] + list(cov_piv.columns))
            for month in MONTH_NAMES:
                row = [month]
                if month in cov_piv.index:
                    row += [float(x) if pd.notna(x) else None for x in cov_piv.loc[month].tolist()]
                else:
                    row += [None] * len(cov_piv.columns)
                ws_cov.append(row)

            ws_cov.freeze_panes = "B2"
            ws_cov.auto_filter.ref = ws_cov.dimensions
            style_header_row(ws_cov, row=1)
            for r in range(2, 2 + len(MONTH_NAMES)):
                for c in range(2, 2 + len(cov_piv.columns)):
                    ws_cov.cell(row=r, column=c).number_format = num_format_pct_1

            # Conditional formatting: red if below threshold
            if cov_piv.shape[1] > 0:
                start_cell = ws_cov.cell(row=2, column=2).coordinate
                end_cell = ws_cov.cell(row=1 + len(MONTH_NAMES), column=1 + cov_piv.shape[1]).coordinate
                ws_cov.conditional_formatting.add(
                    f"{start_cell}:{end_cell}",
                    CellIsRule(
                        operator="lessThan",
                        formula=[str(float(coverage_threshold))],
                        stopIfTrue=True,
                        font=Font(color="9C0006"),
                    ),
                )

            # Coverage summary (by month)
            try:
                below = metrics[(metrics["coverage_pct"].notna()) & (metrics["coverage_pct"] < float(coverage_threshold))].copy()
                by_month = (
                    metrics.groupby("month_name", as_index=False)
                    .agg(
                        mean_coverage=("coverage_pct", "mean"),
                        min_coverage=("coverage_pct", "min"),
                    )
                    .set_index("month_name")
                    .reindex(MONTH_NAMES)
                )
                below_count = (
                    below.groupby("month_name").size().reindex(MONTH_NAMES).fillna(0).astype(int)
                    if not below.empty
                    else pd.Series([0] * len(MONTH_NAMES), index=MONTH_NAMES)
                )
                worst_point = (
                    metrics.sort_values("coverage_pct", ascending=True)
                    .dropna(subset=["coverage_pct"])
                    .groupby("month_name")["point_label"]
                    .first()
                    .reindex(MONTH_NAMES)
                )

                ws_covsum = wb.create_sheet("Coverage summary")
                ws_covsum.append(
                    ["Måned", "Snitt dekning (%)", "Min dekning (%)", "Under terskel (#)", "Verste punkt", "Terskel (%)"]
                )
                style_header_row(ws_covsum, row=1)
                for month in MONTH_NAMES:
                    ws_covsum.append(
                        [
                            month,
                            float(by_month.loc[month, "mean_coverage"])
                            if month in by_month.index and pd.notna(by_month.loc[month, "mean_coverage"])
                            else None,
                            float(by_month.loc[month, "min_coverage"])
                            if month in by_month.index and pd.notna(by_month.loc[month, "min_coverage"])
                            else None,
                            int(below_count.loc[month]) if month in below_count.index else 0,
                            str(worst_point.loc[month]) if month in worst_point.index and pd.notna(worst_point.loc[month]) else None,
                            float(coverage_threshold),
                        ]
                    )

                ws_covsum.freeze_panes = "A2"
                ws_covsum.auto_filter.ref = ws_covsum.dimensions
                ws_covsum.column_dimensions["A"].width = 14
                ws_covsum.column_dimensions["B"].width = 18
                ws_covsum.column_dimensions["C"].width = 16
                ws_covsum.column_dimensions["D"].width = 16
                ws_covsum.column_dimensions["E"].width = 30
                ws_covsum.column_dimensions["F"].width = 12
                for r in range(2, 2 + len(MONTH_NAMES)):
                    ws_covsum.cell(row=r, column=2).number_format = num_format_pct_1
                    ws_covsum.cell(row=r, column=3).number_format = num_format_pct_1
                    ws_covsum.cell(row=r, column=6).number_format = num_format_pct_1

                chart = LineChart()
                chart.title = f"Dekning per måned ({selected_year})"
                chart.y_axis.title = "Dekning (%)"
                chart.x_axis.title = "Måned"
                data_ref = Reference(ws_covsum, min_col=2, min_row=1, max_col=6, max_row=1 + len(MONTH_NAMES))
                cats = Reference(ws_covsum, min_col=1, min_row=2, max_row=1 + len(MONTH_NAMES))
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats)
                chart.height = 9
                chart.width = 20
                ws_covsum.add_chart(chart, "H2")

                if not below.empty:
                    ws_low = wb.create_sheet("Low coverage")
                    ws_low.append(["År", "Måned", "Målepunkt", "Dekning (%)"])
                    style_header_row(ws_low, row=1)
                    below = below.sort_values(["coverage_pct", "month", "point_label"], ascending=[True, True, True])
                    for r in below.itertuples(index=False):
                        ws_low.append(
                            [
                                int(selected_year),
                                str(r.month_name),
                                str(r.point_label),
                                float(r.coverage_pct) if pd.notna(r.coverage_pct) else None,
                            ]
                        )
                    ws_low.freeze_panes = "A2"
                    ws_low.auto_filter.ref = ws_low.dimensions
                    ws_low.column_dimensions["A"].width = 8
                    ws_low.column_dimensions["B"].width = 14
                    ws_low.column_dimensions["C"].width = 30
                    ws_low.column_dimensions["D"].width = 12
                    for rr in range(2, 2 + len(below)):
                        ws_low.cell(row=rr, column=4).number_format = num_format_pct_1
            except Exception as exc:
                logger.warning("Excel: lav-dekning-ark feilet: %s", exc)

            if point == "Ryfast (sum tunneler)":
                point_ids_by_group = {
                    "Ryfylketunnelen": TRAFFIC_POINTS["Ryfylketunnelen"]["ids"],
                    "Hundvågtunnelen": (
                        TRAFFIC_POINTS["Hundvågtunnelen"]["ids"] if include_ramp else HUNDVAG_TUNNEL_IDS_UTEN_PÅRAMPE
                    ),
                }

                grouped = group_coverage_by_month(metrics, point_ids_by_group)
                if not grouped.empty:
                    ws_g = wb.create_sheet("Coverage (tunnel)")
                    ws_g.append(["Måned", "Tunnel", "Dekning (%)"])
                    for row in grouped.itertuples(index=False):
                        ws_g.append([row.month_name, row.group, float(row.coverage_pct) if pd.notna(row.coverage_pct) else None])
                    ws_g.freeze_panes = "A2"
                    ws_g.auto_filter.ref = ws_g.dimensions
                    style_header_row(ws_g, row=1)
                    for rr in range(2, 2 + len(grouped)):
                        ws_g.cell(row=rr, column=3).number_format = num_format_pct_1

                try:
                    totals_by_group_df, coverage_by_group_df = aggregate_monthly_totals_by_group(
                        traffic_by_point, point_ids_by_group, int(selected_year)
                    )
                    if totals_by_group_df is not None and not totals_by_group_df.empty:
                        ws_ttot = wb.create_sheet("Tunnel totals")
                        ws_ttot.append(list(totals_by_group_df.columns))
                        for row in totals_by_group_df.itertuples(index=False):
                            ws_ttot.append(_excel_row(row))
                        ws_ttot.freeze_panes = "A2"
                        ws_ttot.auto_filter.ref = ws_ttot.dimensions
                        style_header_row(ws_ttot, row=1)
                        for col_idx, col_name in enumerate(totals_by_group_df.columns, start=1):
                            if col_name in list(point_ids_by_group.keys()):
                                for rr in range(2, 2 + len(totals_by_group_df)):
                                    ws_ttot.cell(row=rr, column=col_idx).number_format = num_format_int

                        chart = LineChart()
                        chart.title = f"Tunnel-fordeling per måned ({selected_year})"
                        chart.y_axis.title = "Passeringer"
                        chart.x_axis.title = "Måned"
                        data_ref = Reference(
                            ws_ttot,
                            min_col=3,
                            min_row=1,
                            max_col=2 + len(point_ids_by_group),
                            max_row=1 + len(totals_by_group_df),
                        )
                        cats = Reference(ws_ttot, min_col=2, min_row=2, max_row=1 + len(totals_by_group_df))
                        chart.add_data(data_ref, titles_from_data=True)
                        chart.set_categories(cats)
                        ws_ttot.add_chart(chart, "H2")

                    if coverage_by_group_df is not None and not coverage_by_group_df.empty:
                        ws_tcov = wb.create_sheet("Tunnel coverage")
                        ws_tcov.append(list(coverage_by_group_df.columns))
                        for row in coverage_by_group_df.itertuples(index=False):
                            ws_tcov.append(_excel_row(row))
                        ws_tcov.freeze_panes = "A2"
                        ws_tcov.auto_filter.ref = ws_tcov.dimensions
                        style_header_row(ws_tcov, row=1)
                        for col_idx, col_name in enumerate(coverage_by_group_df.columns, start=1):
                            if col_name in list(point_ids_by_group.keys()):
                                for rr in range(2, 2 + len(coverage_by_group_df)):
                                    ws_tcov.cell(row=rr, column=col_idx).number_format = num_format_pct_1
                except Exception as exc:
                    logger.warning("Excel: tunnel-ark feilet: %s", exc)

            # Data quality warnings
            try:
                warnings: List[Tuple[str, str, Optional[str]]] = []

                cov_month = compute_monthly_coverage_summary(traffic_by_point, int(selected_year), point_ids)
                missing_points = cov_month[(cov_month["points_present"] < cov_month["points_expected"]) & (cov_month["points_present"] > 0)]
                for r in missing_points.itertuples(index=False):
                    warnings.append(
                        (
                            "Manglende målepunkter",
                            f"{int(r.year)} {str(r.month_name)}",
                            f"{int(r.points_present)}/{int(r.points_expected)} punkter med data",
                        )
                    )
                none_points = cov_month[cov_month["points_present"] == 0]
                for r in none_points.itertuples(index=False):
                    warnings.append(("Ingen data", f"{int(r.year)} {str(r.month_name)}", None))

                anomalies = detect_monthly_anomalies(df, threshold_pct=float(ANOMALY_THRESHOLD_PCT))
                for r in anomalies.itertuples(index=False):
                    warnings.append(
                        (
                            "Anomali",
                            f"{int(r.year)} {str(r.month_name)}",
                            f"Avvik {float(r.deviation_pct):+.1f}% (faktisk {int(round(float(r.actual))):,} / forventet {int(round(float(r.expected))):,})".replace(
                                ",", " "
                            ),
                        )
                    )

                if warnings:
                    ws_warn = wb.create_sheet("Warnings")
                    ws_warn.append(["Type", "Periode", "Detalj"])
                    style_header_row(ws_warn, row=1)
                    for t, period, detail in warnings:
                        ws_warn.append([t, period, detail])
                    ws_warn.freeze_panes = "A2"
                    ws_warn.auto_filter.ref = ws_warn.dimensions
                    ws_warn.column_dimensions["A"].width = 22
                    ws_warn.column_dimensions["B"].width = 18
                    ws_warn.column_dimensions["C"].width = 70

                    ws_summary["A7"] = f"Varsler: {len(warnings)} (se fanen 'Warnings')"
                    ws_summary["A7"].alignment = wrap_top
            except Exception as exc:
                logger.warning("Excel: varsels-ark feilet: %s", exc)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
